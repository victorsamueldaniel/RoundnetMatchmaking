"""models.py - Player, TeamOfTwo, GameOfFour, GamesRound, SessionOfRounds
plus objective-function helpers."""

import pandas as pd
from pandas.api.types import is_numeric_dtype
import numpy as np
import inspect
from itertools import combinations
from collections import defaultdict
from math import comb
import random
import os
import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Maps spectrum_game dict keys → Player attribute names
_SPEC_KEY_TO_ATTR = {
    "Prey": "prey",
    "Equilibrist": "equilibrist",
    "Challenger": "challenger",
    "Chill": "chill",
    "Hunter": "hunter",
    "Classist": "classist",
}


def _cfg_get(cfg, path, default):
    """Return nested config value from dot path; fallback to default."""
    cur = cfg
    for key in path.split("."):
        if not isinstance(cur, dict) or key not in cur:
            return default
        cur = cur[key]
    return cur


def mean_happiness_objective(self):
    return np.mean([player.happiness for player in self.participants])


def std_happiness_objective(self):
    return np.std([player.happiness for player in self.participants])


def mean_std_happiness_objective(self, lambda_weight=2):
    all_happiness = [player.happiness for player in self.participants]
    happiness_mean = np.mean(all_happiness)
    happiness_std = np.std(all_happiness)
    happiness_score = happiness_mean - lambda_weight * happiness_std
    return happiness_score


def min_happiness_objective(self, percentile=10):
    all_happiness = [player.happiness for player in self.participants]
    bottom_percent = np.percentile(all_happiness, percentile)
    return np.mean(bottom_percent)


def mean_min_max_happiness_objective(self, lambda_weight=2, percentile=10):
    all_happiness = [player.happiness for player in self.participants]
    happiness_mean = np.mean(all_happiness)
    # Get bottom percentile players by happiness
    bottom_percent = np.percentile(all_happiness, percentile)
    # Penalize if any player is below the threshold
    return happiness_mean + lambda_weight * np.mean(
        [h for h in all_happiness if h <= bottom_percent]
    )


def compute_session_score(players, objective_name, lambda_weight=2.4, percentile=10):
    """Compute a scalar session score from a list of Player objects and objective metadata.

    Parameters
    ----------
    players : list of Player
        All players whose happiness to include.
    objective_name : str or None
        One of the _KNOWN_OBJECTIVE_FUNCTIONS names.  Defaults to
        ``mean_min_max_happiness_objective`` when None or unrecognised.
    lambda_weight : float
        Lambda parameter used by mean_std and mean_min_max objectives.
    percentile : float
        Bottom-percentile parameter used by mean_min_max objective.

    Returns
    -------
    float
    """
    all_h = [p.happiness for p in players]
    if not all_h:
        return 0.0

    name = objective_name or "mean_min_max_happiness_objective"

    if name == "mean_happiness_objective":
        return float(np.mean(all_h))

    if name == "std_happiness_objective":
        return float(np.std(all_h))

    if name == "min_happiness_objective":
        pct = percentile if percentile is not None else 10
        return float(np.percentile(all_h, pct))

    if name == "mean_std_happiness_objective":
        lw = lambda_weight if lambda_weight is not None else 2.0
        return float(np.mean(all_h) - lw * np.std(all_h))

    # Default: mean_min_max_happiness_objective
    lw = lambda_weight if lambda_weight is not None else 2.4
    pct = percentile if percentile is not None else 10
    bottom_threshold = np.percentile(all_h, pct)
    bottom_vals = [h for h in all_h if h <= bottom_threshold]
    return float(np.mean(all_h) + lw * np.mean(bottom_vals))


_KNOWN_OBJECTIVE_FUNCTIONS = {
    "mean_happiness_objective",
    "std_happiness_objective",
    "mean_std_happiness_objective",
    "min_happiness_objective",
    "mean_min_max_happiness_objective",
}


def _extract_callable_name(objective_function):
    name = getattr(objective_function, "__name__", None)
    if name and name != "<lambda>":
        return name

    base_callable = getattr(objective_function, "func", None)
    if base_callable is not None:
        base_name = getattr(base_callable, "__name__", None)
        if base_name:
            return base_name

    return name


def _extract_code_names(objective_function):
    names = set()

    code_obj = getattr(objective_function, "__code__", None)
    if code_obj is not None:
        names.update(code_obj.co_names)

    base_callable = getattr(objective_function, "func", None)
    base_code_obj = getattr(base_callable, "__code__", None)
    if base_code_obj is not None:
        names.update(base_code_obj.co_names)

    return names


def _extract_numeric_from_callable(objective_function, key):
    keywords = getattr(objective_function, "keywords", None)
    if isinstance(keywords, dict):
        value = keywords.get(key)
        if isinstance(value, (int, float)):
            return float(value)

    try:
        closure_vars = inspect.getclosurevars(objective_function)
        value = closure_vars.nonlocals.get(key)
        if isinstance(value, (int, float)):
            return float(value)
    except Exception:
        pass

    return None


def _infer_objective_metadata(
    objective_function,
    default_name,
    default_lambda_weight,
    default_percentile=10,
):
    name = _extract_callable_name(objective_function)
    if name not in _KNOWN_OBJECTIVE_FUNCTIONS:
        code_names = _extract_code_names(objective_function)
        for known_name in _KNOWN_OBJECTIVE_FUNCTIONS:
            if known_name in code_names:
                name = known_name
                break

    if name not in _KNOWN_OBJECTIVE_FUNCTIONS:
        name = default_name

    lambda_weight = None
    percentile = None

    if name in {"mean_std_happiness_objective", "mean_min_max_happiness_objective"}:
        lambda_weight = _extract_numeric_from_callable(
            objective_function, "lambda_weight"
        )
        if lambda_weight is None:
            lambda_weight = default_lambda_weight

    if name in {"min_happiness_objective", "mean_min_max_happiness_objective"}:
        percentile = _extract_numeric_from_callable(objective_function, "percentile")
        if percentile is None:
            percentile = default_percentile

    return {
        "name": name,
        "lambda_weight": lambda_weight,
        "percentile": percentile,
    }


def _restore_objective_from_metadata(
    objective_name,
    default_name,
    default_lambda_weight,
    default_percentile,
    lambda_weight=None,
    percentile=None,
):
    name = objective_name or default_name

    if name == "mean_happiness_objective":
        return mean_happiness_objective, name, None, None

    if name == "std_happiness_objective":
        return std_happiness_objective, name, None, None

    if name == "min_happiness_objective":
        resolved_percentile = default_percentile if percentile is None else percentile
        return (
            lambda x: min_happiness_objective(x, percentile=resolved_percentile),
            name,
            None,
            resolved_percentile,
        )

    if name == "mean_std_happiness_objective":
        resolved_lambda_weight = (
            default_lambda_weight if lambda_weight is None else lambda_weight
        )
        return (
            lambda x: mean_std_happiness_objective(
                x, lambda_weight=resolved_lambda_weight
            ),
            name,
            resolved_lambda_weight,
            None,
        )

    if name == "mean_min_max_happiness_objective":
        resolved_lambda_weight = (
            default_lambda_weight if lambda_weight is None else lambda_weight
        )
        resolved_percentile = default_percentile if percentile is None else percentile
        return (
            lambda x: mean_min_max_happiness_objective(
                x,
                lambda_weight=resolved_lambda_weight,
                percentile=resolved_percentile,
            ),
            name,
            resolved_lambda_weight,
            resolved_percentile,
        )

    # Unknown objective names fall back to the configured default objective.
    if default_name == "mean_std_happiness_objective":
        return (
            lambda x: mean_std_happiness_objective(
                x, lambda_weight=default_lambda_weight
            ),
            "mean_std_happiness_objective",
            default_lambda_weight,
            None,
        )

    return (
        lambda x: mean_min_max_happiness_objective(
            x,
            lambda_weight=default_lambda_weight,
            percentile=default_percentile,
        ),
        "mean_min_max_happiness_objective",
        default_lambda_weight,
        default_percentile,
    )


# %%
class Player:
    def __init__(self, series):
        """
        creating Player class from series. The Series should contain at least the following columns:
        - Level : float. The level of the player
        - Gender : str. The gender of the player (currently "Male" and "Female")
        - Happiness: float. The happiness of the player (normally at 0)
        - Games played: int. The number of games played by the player (normally at 0)
        """
        # Initialize the Player object with a pandas Series
        self.series = series
        # Set attributes for each key in the series
        for key in series.keys():
            setattr(self, key.lower().replace(" ", "_"), series[key])
        # Set the name attribute
        self.name = series.name
        # Sets a "noisy level" to randomize a bit the level of the player. Especially useful for games by level
        # the initial noisy level is set to the level of the player, and will change depending on what type of games the player will engage in
        self.noisy_level = series["Level"]
        # we also created a rounded noisy level, to avoid too many categories of levels
        # default is the level of the player
        self.rounded_noisy_level = series["Level"]
        self.happiness = series.get("Happiness", 0)
        self.previous_happiness = self.happiness  # <-- Add this line
        self.relative_happiness = 0  # Will be calculated at end of session

        self.prey = series.get("Prey", 0)
        self.equilibrist = series.get("Equilibrist", 0)
        self.challenger = series.get("Challenger", 0)
        self.chill = series.get("Chill", 0)
        self.hunter = series.get("Hunter", 0)
        self.classist = series.get("Classist", 0)
        self.last_spec_chosen = None
        self.spec_chosen_history = []

        self.last_happiness_gained = 0
        self.happiness_gained_history = []

        self.games_played = series.get("Games played", 0)
        self.teammate_history = []
        self.other_players_in_same_game_history = []

    def update_happiness(
        self,
        game_mean_level,
        teammates_levels,
        opponents_levels,
        level_gap_tol,
        is_gender_preference_satisfied,
        players_chill,
        session_median,
        weight_same_teammate=4,
        has_same_teammate=False,
        amount_same_people_in_game_history=0,
        spectrum=False,
        seed=None,
        type_preference=None,
        gender_preference=None,
        minority_gender=None,
        player_level=None,
        gender_level_medians=None,
        never_met_players_in_game_count=0,
        never_met_bonus_per_player=2,
        never_met_bonus_cap=4,
        preferred_pair_bonus=0,
        spectrum_prey_opponents_mean_level_multiplier=0.7,
        spectrum_challenger_opponents_mean_level_multiplier=0.9,
        spectrum_challenger_level_gap_tol_multiplier=0.5,
        spectrum_equilibrist_level_gap_tol_multiplier=0.5,
        spectrum_classist_level_gap_tol_multiplier=0.5,
        spectrum_chill_players_chill_threshold=10,
        non_spectrum_high_level_threshold_self_level_multiplier=0.85,
        happiness_penalty_same_people_in_game_history_weight_same_teammate_divisor=2,
        happiness_penalty_gender_preference_not_satisfied_spectrum=5,
        happiness_penalty_gender_preference_not_satisfied_non_spectrum=2,
        happiness_bonus_minority_gender_mixed=1,
        happiness_bonus_above_median_level_type_level=1,
    ):
        initial_happiness = self.happiness

        if spectrum:
            teammates_mean_level = np.mean(teammates_levels)
            team_level = np.mean(teammates_levels + [self.level])
            opponents_mean_level = np.mean(opponents_levels)
            spectrum_game = {
                "Prey": (
                    1
                    if spectrum_prey_opponents_mean_level_multiplier
                    * opponents_mean_level
                    >= self.level
                    else 0
                ),
                "Equilibrist": (
                    1
                    if abs(team_level - opponents_mean_level)
                    <= spectrum_equilibrist_level_gap_tol_multiplier * level_gap_tol
                    else 0
                ),
                "Challenger": (
                    1
                    if abs(
                        spectrum_challenger_opponents_mean_level_multiplier
                        * opponents_mean_level
                        - team_level
                    )
                    <= spectrum_challenger_level_gap_tol_multiplier * level_gap_tol
                    else 0
                ),
                "Chill": (
                    1 if players_chill >= spectrum_chill_players_chill_threshold else 0
                ),
                "Hunter": 1 if opponents_mean_level <= team_level else 0,
                "Classist": (
                    1
                    if abs(self.level - teammates_mean_level)
                    <= spectrum_classist_level_gap_tol_multiplier * level_gap_tol
                    else 0
                ),
            }
            best_gain = 0
            specs_with_best_gain = []
            # Sort specs to ensure consistent ordering
            for spec in sorted(spectrum_game.keys()):
                spec_gain = getattr(self, _SPEC_KEY_TO_ATTR[spec]) * spectrum_game[spec]
                if spec_gain > best_gain:
                    specs_with_best_gain = [spec]
                    best_gain = spec_gain
                elif spec_gain == best_gain:
                    specs_with_best_gain.append(spec)

            spec_chosen = random.choice(specs_with_best_gain)
            self.happiness += (
                getattr(self, _SPEC_KEY_TO_ATTR[spec_chosen])
                * spectrum_game[spec_chosen]
            )
            self.last_spec_chosen = spec_chosen

        else:
            # More nuanced happiness calculation for higher level players
            high_level_teammates = sum(
                1
                for level in teammates_levels
                if level
                >= (
                    self.level * non_spectrum_high_level_threshold_self_level_multiplier
                )
            )
            high_level_opponents = sum(
                1
                for level in opponents_levels
                if level
                >= (
                    self.level * non_spectrum_high_level_threshold_self_level_multiplier
                )
            )

            # Higher level players are happier with competitive matches
            self.happiness += high_level_teammates + high_level_opponents

            # Penalize if consistently playing with much lower level players
            # if np.mean(teammates_levels + opponents_levels) < (self.level * 0.85):
            #     self.happiness -= 1
            self.last_spec_chosen = None

        self.happiness -= weight_same_teammate * has_same_teammate

        # Penalize for repeated encounters (same opponents/teammates in a game).
        # Using /2 (instead of /3) gives a stronger nudge toward variety.
        try:
            _same_people_div = float(
                happiness_penalty_same_people_in_game_history_weight_same_teammate_divisor
            )
        except (TypeError, ValueError):
            _same_people_div = 2.0
        if _same_people_div == 0:
            _same_people_div = 1.0
        self.happiness -= (
            weight_same_teammate / _same_people_div * amount_same_people_in_game_history
        )

        # Bonus for meeting players never encountered before (teammates or opponents)
        never_met_bonus = min(
            never_met_players_in_game_count * never_met_bonus_per_player,
            never_met_bonus_cap,
        )
        self.happiness += never_met_bonus

        # Gender preference satisfaction - penalize if game doesn't satisfy preference
        if not is_gender_preference_satisfied:
            self.happiness -= (
                happiness_penalty_gender_preference_not_satisfied_spectrum
                if spectrum
                else happiness_penalty_gender_preference_not_satisfied_non_spectrum
            )

        # Add bonus for minority gender in mixed games
        if gender_preference == "mixed" and minority_gender is not None:
            if self.gender == minority_gender:
                self.happiness += happiness_bonus_minority_gender_mixed

        # Add bonus for above-median level in level-based games
        if type_preference == "level" and player_level is not None:
            # If level game is also mixed, check against gender-specific median
            if gender_preference == "mixed" and gender_level_medians:
                gender_median = gender_level_medians.get(self.gender)
                if gender_median is not None and player_level > gender_median:
                    self.happiness += happiness_bonus_above_median_level_type_level
            # Otherwise check against session median
            elif session_median is not None:
                if player_level > session_median:
                    self.happiness += happiness_bonus_above_median_level_type_level

        self.happiness += preferred_pair_bonus

        # Store the happiness gained for this game
        self.last_happiness_gained = self.happiness - initial_happiness

        return self.last_happiness_gained


# %%
################################################################################
###                                                                         ####
### ####### #######  #####  #     # ####### ####### ####### #     # ####### ####
###    #    #       #     # ##   ## #     # #          #    #     # #     # ####
###    #    ####### ####### #  #  # #     # #####      #    #  #  # #     # ####
###    #    #       #     # #     # #     # #          #    ##   ## #     # ####
###    #    ####### #     # #     # ####### #          #    #     # ####### ####
###                                                                         ####
################################################################################
class TeamOfTwo:
    def __init__(
        self, player_A, player_B, type_preference=None, gender_preference=None
    ):
        # Initialize the TeamOfTwo object with two players and a type_preference
        self.type_preference = type_preference
        self.gender_preference = gender_preference
        self.player_A = player_A
        self.player_B = player_B
        # Store the players in a list to maintain order
        self.players = [player_A, player_B]
        self.players_set = {player_A, player_B}  # Keep set for membership tests
        self.players_frozenset = frozenset(
            self.players_set
        )  # Hashable for O(1) lookups
        self.players_name = {player_A.name, player_B.name}

        # Calculate the level difference between the two players
        self.level_difference = abs(self.player_A.level - self.player_B.level)
        # calculate the mean level of the team
        self.mean_level = (self.player_A.level + self.player_B.level) / 2
        # Determine if the team is mixed gender
        self.mixed = True if self.player_A.gender != self.player_B.gender else False
        # Determine if both players are male
        self.male = (
            True
            if self.player_A.gender == "Male" and self.player_B.gender == "Male"
            else False
        )
        # Determine if both players are female
        self.female = (
            True
            if self.player_A.gender == "Female" and self.player_B.gender == "Female"
            else False
        )
        # Determine if both players are non-binary
        self.non_binary = (
            True
            if self.player_A.gender == "Non-binary"
            and self.player_B.gender == "Non-binary"
            else False
        )

    def same_players(self, other_team):
        # Check if the two teams have the same players
        return self.players_set == other_team.players_set


# %%
###################################################################################
#                                                                                 #
# #######  #####  #     # ####### ####### ####### ####### ####### #     # ######  #
# #       #     # ##   ## #       #     # #       #       #     # #     # #     # #
# #  #### ####### #  #  # ####### #     # #####   #####   #     # #     # ######  #
# #     # #     # #     # #       #     # #       #       #     # #     # #   ##  #
# ####### #     # #     # ####### ####### #       #       ####### ####### #    ## #
#                                                                                 #
###################################################################################
# %%
class GameOfFour:
    def __init__(
        self,
        team_A,
        team_B,
        type_preference=None,
        gender_preference=None,
        weight_same_teammate=4,
    ):
        # Initialize the GameOfFour object with two teams and a type_preference
        self.type_preference = type_preference
        self.gender_preference = gender_preference

        self.team_A = team_A
        self.team_B = team_B

        # Store the teams in a set of frozensets
        self.teams = set([team_A, team_B])
        # Store the participants of the game
        self.participants = frozenset(
            self.team_A.players_set.union(self.team_B.players_set)
        )

        # Calculate and store mean levels
        self.team_A_mean_level = self.team_A.mean_level
        self.team_B_mean_level = self.team_B.mean_level
        self.overall_mean_level = (self.team_A_mean_level + self.team_B_mean_level) / 2

        self.level_difference = np.round(
            abs(self.team_A_mean_level - self.team_B_mean_level), 2
        )
        self.is_gender_preference_satisfied = (
            False if self.compute_gender_preference_score() == 0 else True
        )
        self.weight_same_teammate = weight_same_teammate

    def compute_gender_preference_score(self):
        # return 1 if gender preference is satisfied, 0 otherwise
        if self.gender_preference == "mixed":
            if self.team_A.mixed and self.team_B.mixed:
                return 1
            else:
                return 0
        elif self.gender_preference == "same":
            if (
                (self.team_A.male and self.team_B.male)
                or (self.team_A.female and self.team_B.female)
                or (self.team_A.non_binary and self.team_B.non_binary)
            ):
                return 1
            else:
                return 0
        else:
            return 1

    def update_players_happiness(
        self,
        session_median_level,
        level_gap_tol,
        spectrum,
        seed=None,
        type_preference=None,
        gender_preference=None,
        minority_gender=None,
        gender_level_medians=None,
        never_met_bonus_per_player=2,
        never_met_bonus_cap=4,
        history_cutoff_round_idx=None,
        spectrum_prey_opponents_mean_level_multiplier=0.7,
        spectrum_challenger_opponents_mean_level_multiplier=0.9,
        spectrum_challenger_level_gap_tol_multiplier=0.5,
        spectrum_equilibrist_level_gap_tol_multiplier=0.5,
        spectrum_classist_level_gap_tol_multiplier=0.5,
        spectrum_chill_players_chill_threshold=10,
        non_spectrum_high_level_threshold_self_level_multiplier=0.85,
        happiness_penalty_same_people_in_game_history_weight_same_teammate_divisor=2,
        happiness_penalty_gender_preference_not_satisfied_spectrum=5,
        happiness_penalty_gender_preference_not_satisfied_non_spectrum=2,
        happiness_bonus_minority_gender_mixed=1,
        happiness_bonus_above_median_level_type_level=1,
    ):
        """Update happiness for all players in the game, with penalty for repeated teammates and opponents"""
        for team in [self.team_A, self.team_B]:
            players = list(team.players)
            other_team = self.team_B if team == self.team_A else self.team_A
            opponents_levels_shared = [p.level for p in other_team.players]

            for i, player in enumerate(players):
                teammates_levels = [p.level for j, p in enumerate(players) if j != i]

                if history_cutoff_round_idx is not None:
                    teammate_history = player.teammate_history[
                        :history_cutoff_round_idx
                    ]
                    same_game_history = player.other_players_in_same_game_history[
                        :history_cutoff_round_idx
                    ]
                else:
                    teammate_history = player.teammate_history
                    same_game_history = player.other_players_in_same_game_history

                # Penalty for repeated teammates
                weight_same_teammate = self.weight_same_teammate
                has_same_teammate = any(
                    frozenset([player, teammate]) in teammate_history
                    for teammate in team.players
                )

                # Build met_players once per player from their history
                met_players = set()
                for players_set in same_game_history:
                    met_players.update(players_set)

                # Count repeated opponents
                amount_same_people_in_game_history = sum(
                    1
                    for players_set in same_game_history
                    for other_player in players_set
                    if other_player in self.participants
                )

                never_met_players_in_game_count = sum(
                    1
                    for other_player in self.participants
                    if other_player is not player and other_player not in met_players
                )

                total_players_chill = sum(p.chill for p in players)

                player.update_happiness(
                    game_mean_level=self.overall_mean_level,
                    teammates_levels=teammates_levels,
                    opponents_levels=opponents_levels_shared,
                    level_gap_tol=level_gap_tol,
                    is_gender_preference_satisfied=self.is_gender_preference_satisfied,
                    players_chill=total_players_chill,
                    session_median=session_median_level,
                    weight_same_teammate=weight_same_teammate,
                    has_same_teammate=has_same_teammate,
                    amount_same_people_in_game_history=amount_same_people_in_game_history,
                    spectrum=spectrum,
                    seed=seed,
                    type_preference=type_preference or self.type_preference,
                    gender_preference=gender_preference or self.gender_preference,
                    minority_gender=minority_gender,
                    player_level=player.level,
                    gender_level_medians=gender_level_medians or {},
                    never_met_players_in_game_count=never_met_players_in_game_count,
                    never_met_bonus_per_player=never_met_bonus_per_player,
                    never_met_bonus_cap=never_met_bonus_cap,
                    spectrum_prey_opponents_mean_level_multiplier=spectrum_prey_opponents_mean_level_multiplier,
                    spectrum_challenger_opponents_mean_level_multiplier=spectrum_challenger_opponents_mean_level_multiplier,
                    spectrum_challenger_level_gap_tol_multiplier=spectrum_challenger_level_gap_tol_multiplier,
                    spectrum_equilibrist_level_gap_tol_multiplier=spectrum_equilibrist_level_gap_tol_multiplier,
                    spectrum_classist_level_gap_tol_multiplier=spectrum_classist_level_gap_tol_multiplier,
                    spectrum_chill_players_chill_threshold=spectrum_chill_players_chill_threshold,
                    non_spectrum_high_level_threshold_self_level_multiplier=non_spectrum_high_level_threshold_self_level_multiplier,
                    happiness_penalty_same_people_in_game_history_weight_same_teammate_divisor=happiness_penalty_same_people_in_game_history_weight_same_teammate_divisor,
                    happiness_penalty_gender_preference_not_satisfied_spectrum=happiness_penalty_gender_preference_not_satisfied_spectrum,
                    happiness_penalty_gender_preference_not_satisfied_non_spectrum=happiness_penalty_gender_preference_not_satisfied_non_spectrum,
                    happiness_bonus_minority_gender_mixed=happiness_bonus_minority_gender_mixed,
                    happiness_bonus_above_median_level_type_level=happiness_bonus_above_median_level_type_level,
                )


# %%
###################################################################################
#                                                                                 #
# #######  #####  #     # #######  ###### ######  ####### #     # ##    # #####   #
# #       #     # ##   ## #       #       #     # #     # #     # # #   # #    #  #
# #  #### ####### #  #  # #######  ####   ######  #     # #     # #  #  # #     # #
# #     # #     # #     # #             # #   ##  #     # #     # #   # # #    #  #
# ####### #     # #     # ####### ######  #    ## ####### ####### #    ## #####   #
#                                                                                 #
###################################################################################
# %%
class GamesRound:
    def __init__(
        self,
        list_of_players,
        previous_games_rounds_anti_chron=[],
        teams_per_game=2,
        players_per_team=2,
        amount_of_games=None,
        type_preference=None,
        gender_preference=None,
        num_iter=40,
        level_gap_tol=0.5,
        objective_function=lambda x: mean_std_happiness_objective(x, lambda_weight=2),
        weight_same_teammate=5,
        seed=None,
        spectrum=False,
        minority_gender=None,
        gender_level_medians=None,
        never_met_bonus_per_player=2,
        never_met_bonus_cap=4,
        game_optimization=None,
        happiness=None,
    ):
        if seed is not None:
            random.seed(seed)

        if amount_of_games:
            self.amount_of_games = amount_of_games
        else:
            self.amount_of_games = len(list_of_players) // (
                teams_per_game * players_per_team
            )

        self.type_preference = type_preference
        self.gender_preference = gender_preference
        self.minority_gender = minority_gender
        self.gender_level_medians = gender_level_medians or {}
        self.participants = list_of_players
        self.participants_names = [player.name for player in list_of_players]
        self.previous_games = previous_games_rounds_anti_chron
        self.previous_teams = set(
            team for round in previous_games_rounds_anti_chron for team in round.teams
        )
        # Frozensets of player pairs for O(1) repeated-team detection
        self.previous_team_frozensets = {
            team.players_frozenset for team in self.previous_teams
        }
        self.teams_per_game = teams_per_game
        self.players_per_team = players_per_team
        self.num_iter = num_iter
        self.level_gap_tol = level_gap_tol
        self.spectrum = spectrum
        self.objective_function = objective_function
        self.never_met_bonus_per_player = never_met_bonus_per_player
        self.never_met_bonus_cap = never_met_bonus_cap
        self.game_optimization = game_optimization or {}
        self.happiness_config = happiness or {}
        self._params = self._resolve_params()
        # Create a set of player pairs that have played together
        self.teammate_history = []
        for round in previous_games_rounds_anti_chron:
            for game in round.games:
                for team in game.teams:
                    for player1, player2 in combinations([p for p in team.players], 2):
                        self.teammate_history.append(frozenset([player1, player2]))
        self.games = []
        self.session_median_level = np.median(
            [player.level for player in list_of_players]
        )
        self.weight_same_teammate = weight_same_teammate
        # Initialize iterations attribute to store all game combinations explored
        self.iterations = []
        self.create_games(seed=seed)

        self.not_playing = [
            person for person in list_of_players if person not in self.people_playing
        ]
        for player in self.not_playing:
            player.spec_chosen_history.append(None)
            player.happiness_gained_history.append(None)

    def __getstate__(self):
        """Return state for pickling, excluding non-picklable items"""
        state = self.__dict__.copy()
        objective_metadata = _infer_objective_metadata(
            getattr(self, "objective_function", None),
            default_name="mean_std_happiness_objective",
            default_lambda_weight=2,
            default_percentile=10,
        )
        state["_objective_function_name"] = objective_metadata["name"]
        state["_objective_lambda_weight"] = objective_metadata["lambda_weight"]
        state["_objective_percentile"] = objective_metadata["percentile"]
        # Remove the actual lambda/function as it can't be pickled
        state.pop("objective_function", None)
        return state

    def __setstate__(self, state):
        """Restore state from pickling"""
        func_name = state.pop("_objective_function_name", None)
        lambda_weight = state.pop("_objective_lambda_weight", None)
        percentile = state.pop("_objective_percentile", None)
        self.__dict__.update(state)
        (
            self.objective_function,
            resolved_name,
            resolved_lambda_weight,
            resolved_percentile,
        ) = _restore_objective_from_metadata(
            objective_name=func_name,
            default_name="mean_std_happiness_objective",
            default_lambda_weight=2,
            default_percentile=10,
            lambda_weight=lambda_weight,
            percentile=percentile,
        )
        self._objective_function_name = resolved_name
        self._objective_lambda_weight = resolved_lambda_weight
        self._objective_percentile = resolved_percentile

    def create_set_of_all_possible_teams(self, remove_previous_teams=True):
        # function that creates all possible teams of <players_per_team> players from a set of players
        set_of_all_possible_teams = set(
            TeamOfTwo(*team)
            for team in combinations(self.people_playing, self.players_per_team)
        )
        # remove teams that have already played together (O(1) frozenset lookup)
        if remove_previous_teams:
            set_of_all_possible_teams = {
                team
                for team in set_of_all_possible_teams
                if team.players_frozenset not in self.previous_team_frozensets
            }
        return set_of_all_possible_teams

    def _resolve_params(self):
        """Resolve all config values from game_optimization and happiness_config once."""
        go = self.game_optimization or {}
        hc = self.happiness_config or {}
        return {
            "spectrum_prey_opponents_mean_level_multiplier": _cfg_get(
                go, "spectrum.Prey.opponents_mean_level_multiplier", 0.7
            ),
            "spectrum_challenger_opponents_mean_level_multiplier": _cfg_get(
                go, "spectrum.Challenger.opponents_mean_level_multiplier", 0.9
            ),
            "spectrum_challenger_level_gap_tol_multiplier": _cfg_get(
                go, "spectrum.Challenger.level_gap_tol_multiplier", 0.5
            ),
            "spectrum_equilibrist_level_gap_tol_multiplier": _cfg_get(
                go, "spectrum.Equilibrist.level_gap_tol_multiplier", 0.5
            ),
            "spectrum_classist_level_gap_tol_multiplier": _cfg_get(
                go, "spectrum.Classist.level_gap_tol_multiplier", 0.5
            ),
            "spectrum_chill_players_chill_threshold": _cfg_get(
                go, "spectrum.Chill.players_chill_threshold", 10
            ),
            "non_spectrum_high_level_threshold_self_level_multiplier": _cfg_get(
                go, "non_spectrum.high_level_threshold.self_level_multiplier", 0.85
            ),
            "happiness_penalty_same_people_in_game_history_weight_same_teammate_divisor": _cfg_get(
                hc,
                "penalties.same_people_in_game_history.weight_same_teammate_divisor",
                2,
            ),
            "happiness_penalty_gender_preference_not_satisfied_spectrum": _cfg_get(
                hc, "penalties.gender_preference_not_satisfied.spectrum", 5
            ),
            "happiness_penalty_gender_preference_not_satisfied_non_spectrum": _cfg_get(
                hc, "penalties.gender_preference_not_satisfied.non_spectrum", 2
            ),
            "happiness_bonus_minority_gender_mixed": _cfg_get(
                hc, "bonuses.minority_gender.mixed", 1
            ),
            "happiness_bonus_above_median_level_type_level": _cfg_get(
                hc, "bonuses.above_median_level.type_level", 1
            ),
            "depth_0_cap": max(
                1,
                int(
                    _cfg_get(
                        go, "generate_all_game_combinations.max_combos.depth_0", 20
                    )
                ),
            ),
            "depth_n_cap": max(
                1,
                int(
                    _cfg_get(
                        go, "generate_all_game_combinations.max_combos.depth_n", 10
                    )
                ),
            ),
            "team_combo_cap": max(
                1,
                int(_cfg_get(go, "generate_all_game_combinations.max_team_combos", 3)),
            ),
            "level_sorter_round_factor": _cfg_get(
                go, "games_by_level._level_sorter.round_factor", 1
            ),
            "level_sorter_max_noise_factor": _cfg_get(
                go, "games_by_level._level_sorter.max_noise_factor", 0.2
            ),
        }

    def _happiness_update_kwargs(self, seed, level_gap_tol, spectrum):
        """Return the shared keyword arguments for game.update_players_happiness() calls."""
        p = self._params
        return dict(
            session_median_level=self.session_median_level,
            level_gap_tol=level_gap_tol,
            spectrum=spectrum,
            seed=seed,
            type_preference=self.type_preference,
            gender_preference=self.gender_preference,
            minority_gender=self.minority_gender,
            gender_level_medians=self.gender_level_medians,
            never_met_bonus_per_player=self.never_met_bonus_per_player,
            never_met_bonus_cap=self.never_met_bonus_cap,
            spectrum_prey_opponents_mean_level_multiplier=p[
                "spectrum_prey_opponents_mean_level_multiplier"
            ],
            spectrum_challenger_opponents_mean_level_multiplier=p[
                "spectrum_challenger_opponents_mean_level_multiplier"
            ],
            spectrum_challenger_level_gap_tol_multiplier=p[
                "spectrum_challenger_level_gap_tol_multiplier"
            ],
            spectrum_equilibrist_level_gap_tol_multiplier=p[
                "spectrum_equilibrist_level_gap_tol_multiplier"
            ],
            spectrum_classist_level_gap_tol_multiplier=p[
                "spectrum_classist_level_gap_tol_multiplier"
            ],
            spectrum_chill_players_chill_threshold=p[
                "spectrum_chill_players_chill_threshold"
            ],
            non_spectrum_high_level_threshold_self_level_multiplier=p[
                "non_spectrum_high_level_threshold_self_level_multiplier"
            ],
            happiness_penalty_same_people_in_game_history_weight_same_teammate_divisor=p[
                "happiness_penalty_same_people_in_game_history_weight_same_teammate_divisor"
            ],
            happiness_penalty_gender_preference_not_satisfied_spectrum=p[
                "happiness_penalty_gender_preference_not_satisfied_spectrum"
            ],
            happiness_penalty_gender_preference_not_satisfied_non_spectrum=p[
                "happiness_penalty_gender_preference_not_satisfied_non_spectrum"
            ],
            happiness_bonus_minority_gender_mixed=p[
                "happiness_bonus_minority_gender_mixed"
            ],
            happiness_bonus_above_median_level_type_level=p[
                "happiness_bonus_above_median_level_type_level"
            ],
        )

    def create_games(self, seed=None):

        ####removing players amongst the ones that had played the most##########
        # Calculate how many players will play based on amount_of_games
        max_players_playing = (
            self.amount_of_games * self.players_per_team * self.teams_per_game
        )
        total_players = len(self.participants_names)

        # amount of players that will not play
        if max_players_playing < total_players:
            amount_non_playing = total_players - max_players_playing
        else:
            # If we have fewer players than needed, use all players and adjust by modulo
            amount_non_playing = total_players % (
                self.players_per_team * self.teams_per_game
            )

        # we split the players by amount of games played
        all_amounts_of_games_played = {
            person.games_played for person in self.participants
        }

        dic_amount_of_games_played_list_of_players = {
            amount_of_games_played: []
            for amount_of_games_played in all_amounts_of_games_played
        }
        shuffled_players = self.participants.copy()
        random.shuffle(shuffled_players)

        for player in shuffled_players:
            dic_amount_of_games_played_list_of_players[player.games_played].append(
                player
            )

        # Sort each group by decreasing happiness so, within overplayed players,
        # happier players are benched first and less-happy players keep playing.
        for k in dic_amount_of_games_played_list_of_players:
            dic_amount_of_games_played_list_of_players[k].sort(
                key=lambda p: p.happiness,
                reverse=True,
            )

        list_descending_priority = [
            player
            for i in sorted(
                dic_amount_of_games_played_list_of_players.keys(), reverse=True
            )
            for player in dic_amount_of_games_played_list_of_players[i]
        ]
        self.people_playing = list_descending_priority[amount_non_playing:]

        for player in self.people_playing:
            player.previous_happiness = player.happiness
            player.games_played += 1
        ########################################################################

        self.set_of_all_possible_teams = self.create_set_of_all_possible_teams()
        ######type_preference == none################################################
        # if type_preference is none, we create random games, trying not to recreate the same games
        if isinstance(self.type_preference, dict):
            preference_type = self.type_preference.get("type")
            kwargs = self.type_preference.get("kwargs")
        else:
            preference_type = self.type_preference
            kwargs = {}
        print(
            f"[create_games] preference_type={preference_type}, seed={seed}, type_preference={self.type_preference}"
        )
        if preference_type == "balanced":

            # Create all games together optimizing global happiness score
            self.games = self.create_all_balanced_games(
                self.people_playing,
                level_gap_tol=self.level_gap_tol,
                spectrum=self.spectrum,
                objective_function=self.objective_function,
                seed=seed,
                **kwargs,
            )

            if self.games == []:
                print("could not find a game, because the tolerance is too low")

        if preference_type == "level":
            self.create_games_by_level(seed=seed, **kwargs)
        self.teams = set()
        for game in self.games:
            self.teams = self.teams.union(game.teams)

        # Update history for playing players
        for player in self.people_playing:
            player.spec_chosen_history.append(player.last_spec_chosen)
            player.happiness_gained_history.append(player.last_happiness_gained)

            # Track teammates
            for team in self.teams:
                if player in team.players_set:
                    player.teammate_history.append(frozenset(team.players_set))

            # Track opponents
            for game in self.games:
                if player in game.participants:
                    # Record other players in this game
                    player.other_players_in_same_game_history.append(
                        frozenset([p for p in game.participants if p != player])
                    )

        # Update history for non-playing players (already done in __init__ method)
        # No need to duplicate the history update here

    def create_all_balanced_games(
        self,
        people_playing,
        level_gap_tol,
        spectrum,
        objective_function=lambda x: mean_std_happiness_objective(x, lambda_weight=2),
        seed=None,
        **kwargs,
    ):
        """
        Create all games for a round by considering all possible combinations
        and choosing the one with the best overall happiness score.
        """
        if (
            len(people_playing)
            < self.amount_of_games * self.teams_per_game * self.players_per_team
        ):
            self.iterations = []  # Store empty iterations
            return []

        # Generate all possible ways to divide players into games
        all_game_combinations = self.generate_all_game_combinations(people_playing)

        if not all_game_combinations:
            self.iterations = []  # Store empty iterations
            return []

        best_games = None
        best_overall_score = float("-inf")

        # Two-pass approach: First try with gender preference, then without if needed
        two_pass = self.gender_preference is not None
        iterations_with_scores = []
        best_iteration_data = None

        for pass_num in range(2 if two_pass else 1):
            enforce_gender_preference = (pass_num == 0) if two_pass else False

            # Skip pass 1 if we already found a solution in pass 0
            if pass_num == 1 and best_games is not None:
                break

            # Log when entering pass 1 (relaxes gender constraint)
            # if pass_num == 1:
            #     print(
            #         f"⚠️  Level tolerance too strict with gender constraint - relaxing gender preference..."
            #     )

            for game_combination in all_game_combinations:
                iteration_data = {
                    "games": game_combination,
                    "score": None,
                    "meets_tolerance": False,
                    "meets_level_tolerance": False,
                    "meets_gender_preference": None,
                    "pass_num": pass_num,
                    "gender_enforced": enforce_gender_preference,
                    "selected": False,
                }

                # Check if all games meet level gap tolerance
                meets_level_tolerance = all(
                    game.level_difference <= level_gap_tol for game in game_combination
                )
                iteration_data["meets_level_tolerance"] = meets_level_tolerance
                if not meets_level_tolerance:
                    iterations_with_scores.append(iteration_data)
                    continue

                # Check if all games meet gender preference (on first pass only)
                if enforce_gender_preference:
                    meets_gender_preference = all(
                        game.compute_gender_preference_score() == 1
                        for game in game_combination
                    )
                    iteration_data["meets_gender_preference"] = meets_gender_preference
                    if not meets_gender_preference:
                        iterations_with_scores.append(iteration_data)
                        continue
                else:
                    iteration_data["meets_gender_preference"] = None

                iteration_data["meets_tolerance"] = True

                # Save current happiness state as an indexed list (faster than dict)
                happiness_backup = [
                    (
                        player.happiness,
                        player.previous_happiness,
                        player.last_happiness_gained,
                    )
                    for player in people_playing
                ]

                # Simulate happiness updates for all games
                for game in game_combination:
                    game.update_players_happiness(
                        **self._happiness_update_kwargs(seed, level_gap_tol, spectrum)
                    )

                # Calculate overall happiness score for this combination
                overall_score = objective_function(self)

                # Update iteration data with the score
                iteration_data["score"] = overall_score
                iterations_with_scores.append(iteration_data)

                # Restore happiness state by index (faster than dict lookup)
                for player, (h, ph, lh) in zip(people_playing, happiness_backup):
                    player.happiness = h
                    player.previous_happiness = ph
                    player.last_happiness_gained = lh

                # Track best combination (that meets both tolerance and preference constraints)
                if overall_score > best_overall_score:
                    best_overall_score = overall_score
                    best_games = game_combination
                    best_iteration_data = iteration_data

        if best_iteration_data is not None:
            best_iteration_data["selected"] = True

        # Store all iterations in the instance
        self.iterations = iterations_with_scores

        # Apply the happiness updates for the best combination
        if best_games is not None:
            for game in best_games:
                game.update_players_happiness(
                    **self._happiness_update_kwargs(seed, level_gap_tol, spectrum)
                )

            return list(best_games)
        return []

    def generate_all_game_combinations(self, people_playing):
        """
        Generate all possible ways to divide players into games.
        Returns a list of game combinations, where each combination is a list of GameOfFour objects.
        Uses sampling and early stopping to avoid combinatorial explosion.
        """
        from itertools import combinations

        players = list(people_playing)
        n_players = len(players)
        players_per_game = self.teams_per_game * self.players_per_team

        if n_players != self.amount_of_games * players_per_game:
            return []

        # Strict limit on total combinations to avoid infinite loops
        max_combinations = self.num_iter
        all_combinations = []

        def generate_games_recursive(remaining_players, current_games, depth=0):
            # Early stopping if we have enough combinations
            if len(all_combinations) >= max_combinations:
                return

            if len(remaining_players) == 0:
                # All players assigned, add this combination
                all_combinations.append(current_games.copy())
                return

            if len(remaining_players) < players_per_game:
                # Not enough players for another game
                return

            # Limit the number of player combinations we consider
            game_player_combos = list(combinations(remaining_players, players_per_game))

            # Sample combinations if there are too many
            depth_0_cap = self._params["depth_0_cap"]
            depth_n_cap = self._params["depth_n_cap"]
            if depth == 0:  # First level
                max_combos = min(len(game_player_combos), depth_0_cap)
            else:  # Deeper levels
                max_combos = min(len(game_player_combos), depth_n_cap)

            if len(game_player_combos) > max_combos:
                random.shuffle(game_player_combos)
                game_player_combos = game_player_combos[:max_combos]

            for game_players in game_player_combos:
                if len(all_combinations) >= max_combinations:
                    break

                game_players_list = list(game_players)

                # Limit team arrangements - only try 3 different arrangements
                team_combos = list(
                    combinations(range(players_per_game), self.players_per_team)
                )
                max_team_combos = min(len(team_combos), self._params["team_combo_cap"])
                team_combos = team_combos[:max_team_combos]

                for team1_indices in team_combos:
                    if len(all_combinations) >= max_combinations:
                        break

                    team2_indices = [
                        i for i in range(players_per_game) if i not in team1_indices
                    ]

                    team1_players = [game_players_list[i] for i in team1_indices]
                    team2_players = [game_players_list[i] for i in team2_indices]

                    team1 = TeamOfTwo(*team1_players)
                    team2 = TeamOfTwo(*team2_players)

                    game = GameOfFour(
                        team1,
                        team2,
                        type_preference=self.type_preference,
                        gender_preference=self.gender_preference,
                        weight_same_teammate=self.weight_same_teammate,
                    )

                    # Recursively generate the rest
                    game_players_set = set(game_players)  # O(1) membership
                    new_remaining = [
                        p for p in remaining_players if p not in game_players_set
                    ]
                    current_games.append(game)
                    generate_games_recursive(new_remaining, current_games, depth + 1)
                    current_games.pop()

        generate_games_recursive(players, [])

        return all_combinations

    def _analyze_gender_feasibility(self, players_in_game, gender_preference):
        """
        Analyze if a mixed gender game is possible with given players.
        For a game to be "mixed", BOTH teams must have both genders.

        Returns: (can_be_mixed, gender_counts, reason)
        """
        male_count = sum(1 for p in players_in_game if p.gender == "Male")
        female_count = sum(1 for p in players_in_game if p.gender == "Female")
        gender_counts = {"male": male_count, "female": female_count}

        if gender_preference != "mixed":
            return True, gender_counts, "not_applicable"

        # For a game to be mixed: BOTH teams must have M and F
        # With 4 players split into 2 teams of 2:
        # Only possible if exactly 2M + 2F (balanced)

        if male_count == 2 and female_count == 2:
            return True, gender_counts, "balanced"
        else:
            return False, gender_counts, "imbalanced"

    def _preallocate_players_for_mixed_games(self, people_playing, seed=None):
        """
        Pre-allocate players into groups to maximize mixed games when gender preference is "mixed".

        With M males and F females, calculates max mixed games possible and pre-groups them.

        Returns: list of groups (each group is a list of 4 players to form one game)
        """
        if self.gender_preference != "mixed":
            # No pre-allocation needed
            return None

        male_players = sorted(
            [p for p in people_playing if p.gender == "Male"],
            key=lambda p: self._level_sorter(p, seed=seed),
            reverse=True,
        )
        female_players = sorted(
            [p for p in people_playing if p.gender == "Female"],
            key=lambda p: self._level_sorter(p, seed=seed),
            reverse=True,
        )

        total_players = len(people_playing)
        players_per_game = self.players_per_team * self.teams_per_game
        amount_of_games = total_players // players_per_game

        # Calculate max possible mixed games: each needs 2M + 2F
        max_mixed_games = min(len(male_players) // 2, len(female_players) // 2)

        # Log if we're creating mixed games
        # if max_mixed_games > 0:
        #     print(
        #         f"ℹ️  Organizing for {max_mixed_games} mixed games (M:{len(male_players)}, F:{len(female_players)})"
        #     )

        # Pre-allocate groups
        groups = []

        # Create balanced groups for mixed games
        for i in range(max_mixed_games):
            # Take 2 males and 2 females for each mixed game
            group = [
                male_players[i * 2],
                male_players[i * 2 + 1],
                female_players[i * 2],
                female_players[i * 2 + 1],
            ]
            groups.append(group)

        # Collect remaining players, sorted by level so highest-level group forms first
        remaining_players = sorted(
            male_players[max_mixed_games * 2 :] + female_players[max_mixed_games * 2 :],
            key=lambda p: p.level,
            reverse=True,
        )

        # Form remaining groups from remainder
        for i in range(0, len(remaining_players), players_per_game):
            group = remaining_players[i : i + players_per_game]
            if len(group) == players_per_game:
                groups.append(group)

        return groups if len(groups) == amount_of_games else None

    def _filter_team_pairs_by_gender_preference(
        self, possible_team_pairs_with_level_diff, gender_preference
    ):
        """
        Filter team pairs to only those satisfying gender preference.
        For "mixed": only keep pairs where BOTH teams are mixed (M+F each).
        For "same": only keep pairs where both teams are same gender.
        For "open" or None: keep all pairs.

        Returns list of tuples: (team1, team2, level_diff)
        """
        if gender_preference != "mixed" and gender_preference != "same":
            # No filtering for "open" or None
            return possible_team_pairs_with_level_diff

        filtered_pairs = []
        for team1, team2, level_diff in possible_team_pairs_with_level_diff:
            if gender_preference == "mixed":
                # Both teams must be mixed (have both genders)
                if team1.mixed and team2.mixed:
                    filtered_pairs.append((team1, team2, level_diff))
            elif gender_preference == "same":
                # Both teams must be same gender
                if (
                    (team1.male and team2.male)
                    or (team1.female and team2.female)
                    or (team1.non_binary and team2.non_binary)
                ):
                    filtered_pairs.append((team1, team2, level_diff))

        return filtered_pairs

    def _level_sorter(
        self, player, round_factor=None, max_noise_factor=None, seed=None
    ):
        rng = random.Random(seed)
        if round_factor is None:
            round_factor = self._params["level_sorter_round_factor"]
        try:
            round_factor = float(round_factor)
        except (TypeError, ValueError):
            round_factor = 1.0
        if round_factor == 0:
            round_factor = 1.0
        if max_noise_factor is None:
            max_noise_factor = self._params["level_sorter_max_noise_factor"]
        try:
            max_noise_factor = max(0.0, float(abs(max_noise_factor)))
        except (TypeError, ValueError):
            max_noise_factor = 0.2
        max_noise = self.session_median_level * max_noise_factor
        noisy_level = round(player.level * round_factor) / round_factor + rng.uniform(
            -max_noise, max_noise
        )
        return (noisy_level, -player.happiness)

    def create_games_by_level(self, alternate=False, seed=None, **kwargs):

        # Initialize iterations list to track all possible team arrangements
        iterations_with_scores = []

        # Try to pre-allocate groups for mixed games (if applicable)
        preallocated_groups = self._preallocate_players_for_mixed_games(
            self.people_playing, seed=seed
        )

        if preallocated_groups is not None:
            # Use pre-allocated groups
            all_players_in_each_game = preallocated_groups
        else:
            # Fall back to traditional interleaving/sorting
            if self.gender_preference == "mixed":
                # Separate players by gender
                male_players = sorted(
                    [
                        player
                        for player in self.people_playing
                        if player.gender == "Male"
                    ],
                    key=lambda player: self._level_sorter(player, seed=seed),
                    reverse=True,
                )
                female_players = sorted(
                    [
                        player
                        for player in self.people_playing
                        if player.gender == "Female"
                    ],
                    key=lambda player: self._level_sorter(player, seed=seed),
                    reverse=True,
                )

                # Interleave players to attempt mixed games
                sorted_players = []
                for m, f in zip(male_players, female_players):
                    sorted_players.extend([m, f])
                # Add remaining players if counts don't match
                sorted_players.extend(male_players[len(female_players) :])
                sorted_players.extend(female_players[len(male_players) :])
            else:
                sorted_players = sorted(
                    self.people_playing,
                    key=lambda p: self._level_sorter(p, seed=seed),
                    reverse=True,
                )

            # Shuffle within each level band so different seeds produce different
            # player groupings, increasing variety across rounds.
            if seed is not None:
                rng = random.Random(seed)
                i = 0
                while i < len(sorted_players):
                    band_level = round(sorted_players[i].level)
                    j = i + 1
                    while (
                        j < len(sorted_players)
                        and round(sorted_players[j].level) == band_level
                    ):
                        j += 1
                    band = sorted_players[i:j]
                    rng.shuffle(band)
                    sorted_players[i:j] = band
                    i = j

            all_players_in_each_game = [
                [
                    player
                    for player in sorted_players[
                        i : i + self.players_per_team * self.teams_per_game
                    ]
                ]
                for i in range(
                    0, len(sorted_players), self.players_per_team * self.teams_per_game
                )
            ]

        for game_num, players_in_game in enumerate(all_players_in_each_game):
            start_idx = game_num * self.players_per_team * self.teams_per_game
            iteration_start_idx = len(iterations_with_scores)

            if alternate:
                # Divide players into teams by alternating
                team1_players = players_in_game[0::2][: self.players_per_team]
                team2_players = players_in_game[1::2][: self.players_per_team]
            else:
                # Divide players into teams by putting 0 and 4 together, and 2 and 3 together
                # WORKS ONLY FOR 4 PLAYERS
                if len(players_in_game) == 4:
                    team1_players = [players_in_game[i] for i in [0, 3]]
                    team2_players = [players_in_game[i] for i in [1, 2]]
                else:
                    print(
                        "Error: Game creation by level requires exactly 4 players per game."
                        "please put alternate to True if you want to create games with more than 4 players"
                    )

            # Generate all possible (team1, team2) combinations as TeamOfTwo objects
            alternative_possible_teams = list(
                combinations(players_in_game, self.players_per_team)
            )
            possible_team_pairs_with_level_diff = []
            for team1_players in alternative_possible_teams:
                team2_players = tuple(
                    player for player in players_in_game if player not in team1_players
                )
                if len(team2_players) == self.players_per_team:
                    team1 = TeamOfTwo(*team1_players)
                    team2 = TeamOfTwo(*team2_players)
                    level_diff = abs(team1.mean_level - team2.mean_level)
                    possible_team_pairs_with_level_diff.append(
                        (team1, team2, level_diff)
                    )

            # Sort by level difference (for fallback)
            possible_team_pairs_with_level_diff.sort(key=lambda x: x[2])

            # Analyze gender feasibility and filter if needed
            can_be_mixed, gender_counts, reason = self._analyze_gender_feasibility(
                players_in_game, self.gender_preference
            )

            # Filter pairs by gender preference (if applicable)
            filtered_pairs = self._filter_team_pairs_by_gender_preference(
                possible_team_pairs_with_level_diff, self.gender_preference
            )

            # Log warning if mixed preference but cannot be satisfied
            # if self.gender_preference == "mixed" and not can_be_mixed:
            #     print(
            #         f"⚠️  Game {game_num + 1}: Gender imbalance - cannot make mixed teams (M:{gender_counts['male']}, F:{gender_counts['female']})"
            #     )

            # Store all possible team arrangements as iterations
            for team1, team2, level_diff in possible_team_pairs_with_level_diff:
                iterations_with_scores.append(
                    {
                        "teams": (team1, team2),
                        "level_diff": level_diff,
                        "game_num": game_num,
                        "selected": False,
                    }
                )

            # Pick the first pair matching gender preference AND not in previous_teams.
            found = False

            # Use filtered pairs if any exist, otherwise fall back to all pairs
            pairs_to_check = (
                filtered_pairs
                if filtered_pairs
                else possible_team_pairs_with_level_diff
            )

            for team1, team2, _ in pairs_to_check:
                if (
                    team1.players_frozenset not in self.previous_team_frozensets
                    and team2.players_frozenset not in self.previous_team_frozensets
                ):
                    team1_obj, team2_obj = team1, team2
                    found = True
                    break

            # Final fallback: use first pair from best-effort list
            if not found:
                team1_obj, team2_obj, _ = (
                    pairs_to_check[0]
                    if pairs_to_check
                    else possible_team_pairs_with_level_diff[0]
                )
            team1_players = list(team1_obj.players)
            team2_players = list(team2_obj.players)

            if (
                len(team1_players) == self.players_per_team
                and len(team2_players) == self.players_per_team
            ):
                selected_pair_key = frozenset(
                    (team1_obj.players_frozenset, team2_obj.players_frozenset)
                )
                for iteration_idx in range(
                    iteration_start_idx, len(iterations_with_scores)
                ):
                    iteration_data = iterations_with_scores[iteration_idx]
                    if iteration_data.get("game_num") != game_num:
                        continue
                    iter_team1, iter_team2 = iteration_data["teams"]
                    iter_pair_key = frozenset(
                        (iter_team1.players_frozenset, iter_team2.players_frozenset)
                    )
                    if iter_pair_key == selected_pair_key:
                        iteration_data["selected"] = True
                        break

                team1 = TeamOfTwo(*team1_players)
                team2 = TeamOfTwo(*team2_players)
                game = GameOfFour(
                    team1,
                    team2,
                    type_preference=self.type_preference,
                    gender_preference=self.gender_preference,
                )
                self.games.append(game)
                game.update_players_happiness(
                    **self._happiness_update_kwargs(
                        seed, self.level_gap_tol, self.spectrum
                    )
                )

        # Store all iterations
        self.iterations = iterations_with_scores

    def find_player_position(self, player):
        """Find where a player is located in this round"""
        # Check in games
        for game_idx, game in enumerate(self.games):
            for player_idx, p in enumerate(game.team_A.players):
                if p.name == player.name:
                    return ("game", game_idx, "A", player_idx)
            for player_idx, p in enumerate(game.team_B.players):
                if p.name == player.name:
                    return ("game", game_idx, "B", player_idx)

        # Check in not_playing
        for player_idx, p in enumerate(self.not_playing):
            if p.name == player.name:
                return ("not_playing", player_idx)

        return None

    def swap_player_positions(self, pos1, pos2):
        """Actually swap two players in the data structure"""
        # Get player references
        if pos1[0] == "game":
            _, game_idx1, team_id1, player_idx1 = pos1
            team1 = (
                self.games[game_idx1].team_A
                if team_id1 == "A"
                else self.games[game_idx1].team_B
            )
            player1 = team1.players[player_idx1]
        else:  # not_playing
            _, player_idx1 = pos1
            player1 = self.not_playing[player_idx1]

        if pos2[0] == "game":
            _, game_idx2, team_id2, player_idx2 = pos2
            team2 = (
                self.games[game_idx2].team_A
                if team_id2 == "A"
                else self.games[game_idx2].team_B
            )
            player2 = team2.players[player_idx2]
        else:  # not_playing
            _, player_idx2 = pos2
            player2 = self.not_playing[player_idx2]

        # Perform swap
        if pos1[0] == "game" and pos2[0] == "game":
            # Both in games
            team1.players[player_idx1] = player2
            team2.players[player_idx2] = player1
            team1.mean_level = (team1.players[0].level + team1.players[1].level) / 2
            team2.mean_level = (team2.players[0].level + team2.players[1].level) / 2
        elif pos1[0] == "game" and pos2[0] == "not_playing":
            # Player 1 in game, player 2 not playing
            team1.players[player_idx1] = player2
            self.not_playing[player_idx2] = player1
            team1.mean_level = (team1.players[0].level + team1.players[1].level) / 2
        elif pos1[0] == "not_playing" and pos2[0] == "game":
            # Player 1 not playing, player 2 in game
            self.not_playing[player_idx1] = player2
            team2.players[player_idx2] = player1
            team2.mean_level = (team2.players[0].level + team2.players[1].level) / 2
        else:
            # Both not playing
            self.not_playing[player_idx1] = player2
            self.not_playing[player_idx2] = player1

    def recalculate_happiness(self, round_idx=None):
        """
        Recalculate happiness for all players in this round after player swaps.
        This rebuilds team objects and recalculates all game statistics.

        Parameters:
        -----------
        round_idx : int, optional
            The index of this round in the session (0-based). Used to update history.
        """
        # Save happiness state before recalculation
        happiness_before_this_round = {}
        if round_idx is not None:
            for game in self.games:
                for player in game.participants:
                    # Calculate happiness before this round by subtracting this round's gain
                    if round_idx < len(player.happiness_gained_history):
                        old_happiness_gained = player.happiness_gained_history[
                            round_idx
                        ]
                        if old_happiness_gained is not None:
                            happiness_before_this_round[player.name] = (
                                player.happiness - old_happiness_gained
                            )
                        else:
                            happiness_before_this_round[player.name] = player.happiness
                    else:
                        happiness_before_this_round[player.name] = player.happiness

        # Reset player happiness to before this round
        if happiness_before_this_round:
            for game in self.games:
                for player in game.participants:
                    if player.name in happiness_before_this_round:
                        player.happiness = happiness_before_this_round[player.name]

        # Update team references in games after swaps
        for game in self.games:
            # Rebuild team objects with new player assignments
            team_a_players = game.team_A.players
            team_b_players = game.team_B.players

            # Create new team objects
            new_team_a = TeamOfTwo(
                team_a_players[0],
                team_a_players[1],
                type_preference=game.type_preference,
                gender_preference=game.gender_preference,
            )
            new_team_b = TeamOfTwo(
                team_b_players[0],
                team_b_players[1],
                type_preference=game.type_preference,
                gender_preference=game.gender_preference,
            )

            # Update the game with new teams
            game.team_A = new_team_a
            game.team_B = new_team_b
            game.teams = set([new_team_a, new_team_b])
            game.participants = frozenset(
                new_team_a.players_set.union(new_team_b.players_set)
            )

            # Recalculate team stats
            game.team_A_mean_level = new_team_a.mean_level
            game.team_B_mean_level = new_team_b.mean_level
            game.overall_mean_level = (
                game.team_A_mean_level + game.team_B_mean_level
            ) / 2
            game.level_difference = np.round(
                abs(game.team_A_mean_level - game.team_B_mean_level), 2
            )
            game.is_gender_preference_satisfied = (
                False if game.compute_gender_preference_score() == 0 else True
            )

            # Recalculate happiness for all players in this game
            game.update_players_happiness(
                **self._happiness_update_kwargs(
                    None, self.level_gap_tol, self.spectrum
                ),
                history_cutoff_round_idx=round_idx,
            )

            # Update the happiness history for this round
            if round_idx is not None:
                for player in game.participants:
                    if round_idx < len(player.happiness_gained_history):
                        player.happiness_gained_history[round_idx] = (
                            player.last_happiness_gained
                        )

                    # Update spec_chosen_history for this round
                    if round_idx < len(player.spec_chosen_history):
                        player.spec_chosen_history[round_idx] = player.last_spec_chosen

                    # Update teammate history for this round
                    for team in [game.team_A, game.team_B]:
                        if player in team.players_set:
                            if round_idx < len(player.teammate_history):
                                player.teammate_history[round_idx] = frozenset(
                                    team.players_set
                                )

                    # Update other players in same game history for this round
                    if round_idx < len(player.other_players_in_same_game_history):
                        player.other_players_in_same_game_history[round_idx] = (
                            frozenset([p for p in game.participants if p != player])
                        )

        # Update not_playing players' history (they don't gain happiness).
        # Symmetrically roll back any happiness gained in a previous version of
        # this round before clearing the round-local state, so a swap-and-swap-back
        # returns to exactly the same score as before the first apply.
        if round_idx is not None:
            for player in self.not_playing:
                if round_idx < len(player.happiness_gained_history):
                    player.happiness_gained_history[round_idx] = None
                if round_idx < len(player.spec_chosen_history):
                    player.spec_chosen_history[round_idx] = None
                # Clear stale played-round teammate / opponent history so downstream
                # happiness re-computations don't count a benched round as a meeting.
                if round_idx < len(player.teammate_history):
                    player.teammate_history[round_idx] = frozenset()
                if round_idx < len(player.other_players_in_same_game_history):
                    player.other_players_in_same_game_history[round_idx] = frozenset()

        # Update the round's teams set with all current teams
        self.teams = set()
        for game in self.games:
            self.teams = self.teams.union(game.teams)


# %%
if __name__ == "__main__":
    import data_loader
    import core.models

    list_of_players = [
        core.models.Player(data_loader.main_df.loc[name])
        for name in data_loader.main_df.iloc[10:21].index
    ]
    gender_level_medians = {
        "Male": np.median(
            [
                data_loader.main_df.loc[name]["Level"]
                for name in data_loader.main_df.iloc[10:21].index
                if data_loader.main_df.loc[name]["Gender"] == "Male"
            ]
        ),
        "Female": np.median(
            [
                data_loader.main_df.loc[name]["Level"]
                for name in data_loader.main_df.iloc[10:21].index
                if data_loader.main_df.loc[name]["Gender"] == "Female"
            ]
        ),
    }
    round_of_games = core.models.GamesRound(
        list_of_players,
        type_preference="balanced",
        gender_preference=None,
        gender_level_medians=gender_level_medians,
        level_gap_tol=2,
        num_iter=40,
        seed=1,
    )

    for attr in ["amount_of_games", "type_preference"]:
        print(attr + " : ", getattr(round_of_games, attr))
    print("not playing:", [player.name for player in round_of_games.not_playing])
    i = 1
    for game in round_of_games.games:
        print("_______", "game", i, "_______")
        print([team.players_name for team in game.teams])
        for attr in ["level_difference"]:
            print(attr + " : ", getattr(game, attr))
        i += 1

# %%
###########################################################################################
#                                                                                         #
#  ###### ####### ####### ####### ####### ######  ####### #     # ##    # #####    ###### #
# #       #          #    #     # #       #     # #     # #     # # #   # #    #  #       #
#  ####   #######    #    #     # #####   ######  #     # #     # #  #  # #     #  ####   #
#       # #          #    #     # #       #   ##  #     # #     # #   # # #    #        # #
# ######  #######    #    ####### #       #    ## ####### ####### #    ## #####   ######  #
#                                                                                         #
###########################################################################################


class SessionOfRounds:
    __module__ = "main"  # Explicitly set module name for pickle

    def __init__(
        self,
        list_of_players,
        amount_of_rounds=4,
        games_per_round_each_round=None,
        players_per_team_each_round=None,
        type_preferences=["balanced", "balanced", "level", "level"],
        gender_preferences=["open", "mixed", "mixed", "open"],
        rounds_reordering=None,
        level_gap_tol=1.1,
        num_iter=435,
        spectrum=True,
        objective_function=lambda x: mean_min_max_happiness_objective(
            x, lambda_weight=2.4
        ),
        weight_same_teammate=5,
        never_met_bonus_per_player=2,
        never_met_bonus_cap=4,
        extra_parameters=None,
        prioritize_level_rounds=True,
        seed=None,
    ):
        self.amount_of_rounds = amount_of_rounds
        self.games_per_round_each_round = games_per_round_each_round
        self.players_per_team_each_round = players_per_team_each_round
        self.type_preferences = type_preferences
        self.gender_preferences = gender_preferences
        self.level_gap_tol = level_gap_tol
        self.num_iter = num_iter

        self.players = list_of_players
        self.players_name = [player.name for player in list_of_players]
        self.rounds = []
        self.spectrum = spectrum
        self.objective_function = objective_function
        self.weight_same_teammate = weight_same_teammate
        self.never_met_bonus_per_player = never_met_bonus_per_player
        self.never_met_bonus_cap = never_met_bonus_cap
        self.extra_parameters = extra_parameters or {}
        self.game_optimization = self.extra_parameters.get("game_optimization", {})
        self.happiness_config = self.extra_parameters.get("happiness", {})
        self.prioritize_level_rounds = prioritize_level_rounds
        self.rounds_reordering = rounds_reordering

        # reformatting type_preferences to the amount of type_preferences wanted
        if type_preferences is None:
            type_preferences = [None] * amount_of_rounds
        elif isinstance(type_preferences, str):
            type_preferences = [type_preferences] * amount_of_rounds
        elif len(type_preferences) < amount_of_rounds:
            type_preferences = type_preferences + [None] * (
                amount_of_rounds - len(type_preferences)
            )
        elif len(type_preferences) > amount_of_rounds:
            type_preferences = type_preferences[:amount_of_rounds]
        self.type_preferences = type_preferences

        # reformatting gender_preferences to the amount of gender_preferences wanted
        if gender_preferences is None:
            gender_preferences = [None] * amount_of_rounds
        elif isinstance(gender_preferences, str):
            gender_preferences = [gender_preferences] * amount_of_rounds
        elif len(gender_preferences) < amount_of_rounds:
            gender_preferences = gender_preferences + [None] * (
                amount_of_rounds - len(gender_preferences)
            )
        elif len(gender_preferences) > amount_of_rounds:
            gender_preferences = gender_preferences[:amount_of_rounds]
        self.gender_preferences = gender_preferences

        # reformatting games_per_round_at_each_round to the amount of rounds wanted
        if games_per_round_each_round is None:
            maximal_games_per_round = len(list_of_players) // 4
            games_per_round_each_round = [maximal_games_per_round] * amount_of_rounds
        elif isinstance(games_per_round_each_round, int):
            games_per_round_each_round = [games_per_round_each_round] * amount_of_rounds
        elif len(games_per_round_each_round) < amount_of_rounds:
            games_per_round_each_round = games_per_round_each_round + [
                games_per_round_each_round[-1]
            ] * (amount_of_rounds - len(games_per_round_each_round))

        self.games_per_round_each_round = games_per_round_each_round

        # reformatting players_per_team_each_round to the amount of rounds wanted
        if players_per_team_each_round is None:
            players_per_team_each_round = [2] * self.amount_of_rounds

        elif isinstance(players_per_team_each_round, int):
            players_per_team_each_round = [
                players_per_team_each_round
            ] * amount_of_rounds
        elif len(players_per_team_each_round) < amount_of_rounds:
            players_per_team_each_round = players_per_team_each_round + [2] * (
                amount_of_rounds - len(players_per_team_each_round)
            )

        self.players_per_team_each_round = players_per_team_each_round

        # Calculate majority and minority genders
        self._calculate_gender_distribution()

        self.create_rounds(seed=seed)

    def __getstate__(self):
        """Return state for pickling, excluding non-picklable items"""
        state = self.__dict__.copy()
        objective_metadata = _infer_objective_metadata(
            getattr(self, "objective_function", None),
            default_name="mean_min_max_happiness_objective",
            default_lambda_weight=2.4,
            default_percentile=10,
        )
        state["_objective_function_name"] = objective_metadata["name"]
        state["_objective_lambda_weight"] = objective_metadata["lambda_weight"]
        state["_objective_percentile"] = objective_metadata["percentile"]
        # Remove the actual lambda/function as it can't be pickled
        state.pop("objective_function", None)
        return state

    def __setstate__(self, state):
        """Restore state from pickling"""
        func_name = state.pop("_objective_function_name", None)
        lambda_weight = state.pop("_objective_lambda_weight", None)
        percentile = state.pop("_objective_percentile", None)

        # Legacy compatibility: historical std-marker restores std with lambda=2.
        if func_name == "mean_std_happiness_objective" and lambda_weight is None:
            lambda_weight = 2

        self.__dict__.update(state)

        (
            self.objective_function,
            resolved_name,
            resolved_lambda_weight,
            resolved_percentile,
        ) = _restore_objective_from_metadata(
            objective_name=func_name,
            default_name="mean_min_max_happiness_objective",
            default_lambda_weight=2.4,
            default_percentile=10,
            lambda_weight=lambda_weight,
            percentile=percentile,
        )
        self._objective_function_name = resolved_name
        self._objective_lambda_weight = resolved_lambda_weight
        self._objective_percentile = resolved_percentile

    def _calculate_gender_distribution(self):
        """Calculate the majority and minority genders in the session, and gender-specific level medians"""
        gender_counts = {}
        gender_level_groups = {}

        for player in self.players:
            gender = player.gender
            gender_counts[gender] = gender_counts.get(gender, 0) + 1

            # Collect levels by gender
            if gender not in gender_level_groups:
                gender_level_groups[gender] = []
            gender_level_groups[gender].append(player.level)

        # Sort by count to identify majority and minority
        sorted_genders = sorted(gender_counts.items(), key=lambda x: x[1], reverse=True)

        if len(sorted_genders) >= 2:
            self.majority_gender = sorted_genders[0][0]
            self.minority_gender = sorted_genders[1][0]
        elif len(sorted_genders) == 1:
            self.majority_gender = sorted_genders[0][0]
            self.minority_gender = None
        else:
            self.majority_gender = None
            self.minority_gender = None

        # Calculate gender-specific level medians
        self.gender_level_medians = {}
        for gender, levels in gender_level_groups.items():
            if levels:
                self.gender_level_medians[gender] = np.median(levels)

    def reorder_rounds(self, order_num_list):
        """
        Reorder rounds based on the provided order_num_list.
        Also reorders all player histories and session-level round-indexed attributes to maintain consistency.

        Args:
            order_num_list: List of 1-indexed round numbers in the desired order
        """
        # Reorder rounds based on the provided order_num_list
        self.rounds = [self.rounds[i - 1] for i in order_num_list]

        # Reorder session-level round-indexed attributes
        if (
            hasattr(self, "games_per_round_each_round")
            and self.games_per_round_each_round
        ):
            self.games_per_round_each_round = [
                self.games_per_round_each_round[i - 1] for i in order_num_list
            ]

        if (
            hasattr(self, "players_per_team_each_round")
            and self.players_per_team_each_round
        ):
            self.players_per_team_each_round = [
                self.players_per_team_each_round[i - 1] for i in order_num_list
            ]

        if hasattr(self, "type_preferences") and self.type_preferences:
            self.type_preferences = [
                self.type_preferences[i - 1] for i in order_num_list
            ]

        if hasattr(self, "gender_preferences") and self.gender_preferences:
            self.gender_preferences = [
                self.gender_preferences[i - 1] for i in order_num_list
            ]

        # Reorder all player histories to match the new round order
        for player in self.players:
            # Reorder spec_chosen_history
            if hasattr(player, "spec_chosen_history") and len(
                player.spec_chosen_history
            ) >= len(order_num_list):
                player.spec_chosen_history = [
                    player.spec_chosen_history[i - 1] for i in order_num_list
                ]

            # Reorder happiness_gained_history
            if hasattr(player, "happiness_gained_history") and len(
                player.happiness_gained_history
            ) >= len(order_num_list):
                player.happiness_gained_history = [
                    player.happiness_gained_history[i - 1] for i in order_num_list
                ]

            # Reorder teammate_history
            if hasattr(player, "teammate_history") and len(
                player.teammate_history
            ) >= len(order_num_list):
                player.teammate_history = [
                    player.teammate_history[i - 1] for i in order_num_list
                ]

            # Reorder other_players_in_same_game_history
            if hasattr(player, "other_players_in_same_game_history") and len(
                player.other_players_in_same_game_history
            ) >= len(order_num_list):
                player.other_players_in_same_game_history = [
                    player.other_players_in_same_game_history[i - 1]
                    for i in order_num_list
                ]

        # Rebuild all happiness from scratch in display order so that the
        # initial session score and Games-Editor recalculations share the
        # same chronological context (displayed rounds are authoritative).
        num_rounds = len(self.rounds)
        for player in self.players:
            player.happiness = 0
            player.happiness_gained_history = [None] * num_rounds
            player.teammate_history = [frozenset()] * num_rounds
            player.other_players_in_same_game_history = [frozenset()] * num_rounds
            player.spec_chosen_history = [None] * num_rounds
        for r_idx, r in enumerate(self.rounds):
            r.recalculate_happiness(r_idx)

    def create_rounds(self, seed=None):
        # function that creates a list of rounds
        if self.prioritize_level_rounds:

            def _round_type_priority(type_pref):
                if isinstance(type_pref, dict):
                    round_type = type_pref.get("type")
                else:
                    round_type = type_pref

                round_type = round_type.lower() if isinstance(round_type, str) else None

                if round_type == "level":
                    return 0
                if round_type == "balanced":
                    return 1
                return 2

            ordered_indices = sorted(
                range(self.amount_of_rounds),
                key=lambda idx: (_round_type_priority(self.type_preferences[idx]), idx),
            )

            self.type_preferences = [self.type_preferences[i] for i in ordered_indices]
            self.gender_preferences = [
                self.gender_preferences[i] for i in ordered_indices
            ]
            self.games_per_round_each_round = [
                self.games_per_round_each_round[i] for i in ordered_indices
            ]
            self.players_per_team_each_round = [
                self.players_per_team_each_round[i] for i in ordered_indices
            ]

        rounds = []
        for i in range(self.amount_of_rounds):
            if seed is not None:
                round_seed = seed + i
            else:
                round_seed = None
            rounds.append(
                GamesRound(
                    list_of_players=self.players,
                    amount_of_games=self.games_per_round_each_round[i],
                    players_per_team=self.players_per_team_each_round[i],
                    previous_games_rounds_anti_chron=rounds,
                    type_preference=self.type_preferences[i],
                    gender_preference=self.gender_preferences[i],
                    level_gap_tol=self.level_gap_tol,
                    num_iter=self.num_iter,
                    objective_function=self.objective_function,
                    weight_same_teammate=self.weight_same_teammate,
                    seed=round_seed,
                    spectrum=self.spectrum,
                    minority_gender=self.minority_gender,
                    gender_level_medians=self.gender_level_medians,
                    never_met_bonus_per_player=self.never_met_bonus_per_player,
                    never_met_bonus_cap=self.never_met_bonus_cap,
                    game_optimization=self.game_optimization,
                    happiness=self.happiness_config,
                )
            )
        # Calculate happiness inequality before each round
        for i in range(self.amount_of_rounds):
            # Sort players by happiness to prioritize less happy players
            sorted_players = sorted(self.players, key=lambda p: p.happiness)

            # Give priority to players with lower happiness scores
            if i > 0:  # Skip for first round since all start at 0 happiness
                # Assign temporary boost to level for less happy players
                for idx, player in enumerate(sorted_players):
                    boost_factor = 1 + (
                        0.2 * (len(sorted_players) - idx) / len(sorted_players)
                    )
                    player.temp_boost = boost_factor

            # Reset temporary boosts
            for player in self.players:
                player.temp_boost = 1.0
        self.rounds = rounds
        self.mean_happiness = np.mean([player.happiness for player in self.players])
        self.max_and_min_happiness = (
            np.max([player.happiness for player in self.players]),
            np.min([player.happiness for player in self.players]),
        )
        self.max_happiness_difference = (
            self.max_and_min_happiness[0] - self.max_and_min_happiness[1]
        )
        self.std_happiness = np.std([player.happiness for player in self.players])

        # Calculate relative happiness (centered around mean) for each player
        for player in self.players:
            player.relative_happiness = player.happiness - self.mean_happiness

        if self.rounds_reordering is not None:
            self.reorder_rounds(self.rounds_reordering)
        self.least_happy_players = sorted(self.players, key=lambda p: p.happiness)[:3]

        self.happiest_players = sorted(
            self.players, key=lambda p: p.happiness, reverse=True
        )[:3]

    # %%
    ################################################################################
    #####                                                                      #####
    ##### ####### ######  ####### ##    # #######       #####  #     # #     # #####
    ##### #     # #     #    #    # #   #    #         #     # #     #  #   #  #####
    ##### ####### ######     #    #  #  #    #         ####### #     #    #    #####
    ##### #       #   ##     #    #   # #    #         #     # #     #  #   #  #####
    ##### #       #    ## ####### #    ##    #         #     # ####### #     # #####
    #####                                                                      #####
    ################################################################################

    def count_all_pairs(self, order_num_list=None):
        # Find and display players who played at least twice with each other and record the rounds
        player_pairs = {}
        pair_rounds = {}
        # Use the possibly reordered rounds for analysis
        round_copy = self.rounds.copy()
        if order_num_list is not None:
            round_copy = [round_copy[i - 1] for i in order_num_list]
        for round_index, round in enumerate(round_copy, start=1):
            for game in round.games:
                for team in game.teams:
                    for player_a, player_b in combinations(team.players, 2):
                        pair = frozenset([player_a.name, player_b.name])
                        player_pairs[pair] = player_pairs.get(pair, 0) + 1
                        if pair not in pair_rounds:
                            pair_rounds[pair] = []
                        pair_rounds[pair].append(round_index)
        return player_pairs, pair_rounds

    def count_all_opponent_pairs(self, order_num_list=None):
        # Find and display players who played against each other at least twice and record the rounds
        opponent_pairs = {}
        opponent_pair_rounds = {}
        # Use the possibly reordered rounds for analysis
        round_copy = self.rounds.copy()
        if order_num_list is not None:
            round_copy = [round_copy[i - 1] for i in order_num_list]
        for round_index, round in enumerate(round_copy, start=1):
            for game in round.games:
                team_A, team_B = list(game.teams)
                for player_a in team_A.players:
                    for player_b in team_B.players:
                        pair = frozenset([player_a.name, player_b.name])
                        opponent_pairs[pair] = opponent_pairs.get(pair, 0) + 1
                        if pair not in opponent_pair_rounds:
                            opponent_pair_rounds[pair] = []
                        opponent_pair_rounds[pair].append(round_index)
        return opponent_pairs, opponent_pair_rounds

    def add_team_repetition_to_output(
        self, output, player_pairs, pair_rounds, minimum=2
    ):
        teams_to_repetitions = {}
        output.append(
            f"\n#######PLAYERS WHO PLAYED TOGETHER AT LEAST {minimum} TIMES #######"
        )
        for pair, count in player_pairs.items():
            if count >= minimum:
                rounds = ", ".join(map(str, pair_rounds[pair]))
                output.append(
                    f"{', '.join(pair)} played together {count} times in rounds: {rounds}"
                )
                teams_to_repetitions[pair] = count
        return output, teams_to_repetitions

    def add_opponent_team_repetition_to_output(
        self, output, opponent_pairs, opponent_pair_rounds, minimum=3
    ):
        teams_to_repetitions = {}
        output.append(
            f"\n#######PLAYERS WHO PLAYED AGAINST EACH OTHER AT LEAST {minimum} TIMES #######"
        )
        for pair, count in opponent_pairs.items():
            if count >= minimum:
                rounds = ", ".join(map(str, opponent_pair_rounds[pair]))
                output.append(
                    f"{', '.join(pair)} played against each other {count} times in rounds: {rounds}"
                )
                teams_to_repetitions[pair] = count
        return output, teams_to_repetitions

    def add_never_met_pairs_to_output(self, output, order_num_list=None):
        never_met_pairs = self.get_never_met_pairs(order_num_list)
        never_met_sentences = self.get_never_met_sentences(order_num_list)

        output.append("\n#######PLAYERS THAT NEVER MET#######")
        if never_met_sentences:
            for sentence in never_met_sentences:
                output.append(sentence)
        else:
            output.append("None")

        return output, never_met_pairs

    def get_never_met_pairs(self, order_num_list=None):
        all_player_name_pairs = {
            frozenset([player_a.name, player_b.name])
            for player_a, player_b in combinations(self.players, 2)
        }
        player_pairs, _ = self.count_all_pairs(order_num_list)
        opponent_pairs, _ = self.count_all_opponent_pairs(order_num_list)
        met_pairs = set(player_pairs.keys()).union(set(opponent_pairs.keys()))

        return sorted(
            [pair for pair in all_player_name_pairs if pair not in met_pairs],
            key=lambda pair: sorted(list(pair)),
        )

    def get_never_met_sentences(self, order_num_list=None):
        never_met_pairs = self.get_never_met_pairs(order_num_list)
        never_met_by_player = defaultdict(set)

        for pair in never_met_pairs:
            player_a, player_b = sorted(list(pair))
            never_met_by_player[player_a].add(player_b)
            never_met_by_player[player_b].add(player_a)

        sentences = []
        for player_name in sorted(never_met_by_player.keys()):
            others = sorted(never_met_by_player[player_name])
            if others:
                sentences.append(
                    f"{player_name} never met {', '.join(str(o) for o in others)}"
                )

        return sentences

    # %%
    ################################################################################
    #                                                                              #
    # ####### ######  ####### ##    # #######      #     #  #####  ####### ##    # #
    # #     # #     #    #    # #   #    #         ##   ## #     #    #    # #   # #
    # ####### ######     #    #  #  #    #         #  #  # #######    #    #  #  # #
    # #       #   ##     #    #   # #    #         #     # #     #    #    #   # # #
    # #       #    ## ####### #    ##    #         #     # #     # ####### #    ## #
    #                                                                              #
    ################################################################################

    def print_all_results(
        self,
        print_levels=True,
        order_num_list=None,
        minimum_team_repeats=2,
        minimum_opponent_repeats=3,
    ):
        # ANSI color codes
        class Colors:
            HEADER = "\033[95m"
            OKBLUE = "\033[94m"
            OKCYAN = "\033[96m"
            OKGREEN = "\033[92m"
            WARNING = "\033[93m"
            FAIL = "\033[91m"
            ENDC = "\033[0m"
            BOLD = "\033[1m"
            UNDERLINE = "\033[4m"
            MAGENTA = "\033[35m"
            YELLOW = "\033[33m"
            RED = "\033[31m"
            GREEN = "\033[32m"
            BLUE = "\033[34m"
            CYAN = "\033[36m"

        # Spectrum colors mapping
        spec_colors = {
            "Prey": Colors.RED,
            "Equilibrist": Colors.GREEN,
            "Challenger": Colors.YELLOW,
            "Chill": Colors.CYAN,
            "Hunter": Colors.MAGENTA,
            "Classist": Colors.BLUE,
            None: Colors.ENDC,
        }

        # Collect all printed information
        output = []
        colored_output = []  # For terminal display with colors

        header_text = "all players: " + str(self.players_name)
        output.append(header_text)
        colored_output.append(Colors.HEADER + Colors.BOLD + header_text + Colors.ENDC)

        i = 1
        round_copy = self.rounds.copy()
        if order_num_list is not None:
            round_copy = [round_copy[i - 1] for i in order_num_list]
        for round in round_copy:
            output.append("\n")
            colored_output.append("\n")

            round_header = f"{i} " * 8 + "ROUND " + f"{i} " * 8
            output.append(round_header)
            colored_output.append(
                Colors.OKBLUE + Colors.BOLD + round_header + Colors.ENDC
            )

            pref_text = "type preference : " + str(getattr(round, "type_preference"))
            pref_text += ",\ngender preference : " + str(
                getattr(round, "gender_preference")
            )
            output.append(pref_text)
            colored_output.append(Colors.OKCYAN + pref_text + Colors.ENDC)

            not_playing_text = "not playing: " + str(
                [player.name for player in round.not_playing]
            )
            output.append(not_playing_text)
            colored_output.append(Colors.WARNING + not_playing_text + Colors.ENDC)

            j = 1
            for game in round.games:
                game_header = f"------- game {j} -------"
                output.append(game_header)
                colored_output.append(Colors.OKGREEN + game_header + Colors.ENDC)

                teams_text = str([team.players_name for team in game.teams])
                output.append(teams_text)
                colored_output.append(Colors.BOLD + teams_text + Colors.ENDC)

                # Always print happiness gained for all players
                for player in game.participants:
                    # Get the correct round index for spec_chosen_history
                    round_idx = (
                        order_num_list[i - 1] - 1
                        if order_num_list is not None
                        else i - 1
                    )
                    spec_chosen = player.spec_chosen_history[round_idx]
                    if spec_chosen is None:
                        spec_chosen = ""
                    else:
                        spec_chosen = "(" + spec_chosen + ")"
                    happiness_gained = player.happiness_gained_history[round_idx]

                    if print_levels:
                        # Color code based on happiness gained
                        color = ""
                        if happiness_gained is None:
                            color = Colors.ENDC
                        else:
                            if happiness_gained >= 3:
                                color = (
                                    spec_colors.get(spec_chosen, Colors.ENDC)
                                    + Colors.GREEN
                                )
                            elif 0 <= happiness_gained < 3:
                                color = (
                                    spec_colors.get(spec_chosen, Colors.ENDC)
                                    + Colors.YELLOW
                                )
                            else:  # negative
                                color = (
                                    spec_colors.get(spec_chosen, Colors.ENDC)
                                    + Colors.RED
                                )

                        colored_player_text = f"name : {player.name}, level : {player.level}, happiness gained {spec_chosen}: {color}{np.round(happiness_gained, 2)}{Colors.ENDC}"
                        player_text = f"name : {player.name}, level : {player.level}, happiness gained {spec_chosen}: {np.round(happiness_gained, 2)}"
                    else:
                        # Color code based on happiness gained
                        color = ""
                        if happiness_gained is None:
                            color = Colors.ENDC
                        else:
                            if happiness_gained >= 3:
                                color = (
                                    spec_colors.get(spec_chosen, Colors.ENDC)
                                    + Colors.GREEN
                                )
                            elif 0 <= happiness_gained < 3:
                                color = (
                                    spec_colors.get(spec_chosen, Colors.ENDC)
                                    + Colors.YELLOW
                                )
                            else:  # negative
                                color = (
                                    spec_colors.get(spec_chosen, Colors.ENDC)
                                    + Colors.RED
                                )
                        colored_player_text = f"name : {player.name}, happiness gained {spec_chosen}: {color}{np.round(happiness_gained, 2)}{Colors.ENDC}"
                        player_text = f"name : {player.name}, happiness gained {spec_chosen}: {np.round(happiness_gained, 2)}"

                    output.append(player_text)
                    colored_output.append(colored_player_text)

                if game.type_preference == "balanced":
                    level_diff_text = "level_difference : " + str(
                        np.round(getattr(game, "level_difference"), 2)
                    )
                    output.append(level_diff_text)
                    colored_output.append(level_diff_text)

                if game.type_preference == "level":
                    level_diff_text = (
                        f"level difference : {np.round(game.level_difference, 2)}"
                    )
                    output.append(level_diff_text)
                    colored_output.append(level_diff_text)
                j += 1
            round_end_text = f"{i} " * 8 + "ROUND END " + f"{i} " * 8
            output.append(round_end_text)
            colored_output.append(
                Colors.OKBLUE + Colors.BOLD + round_end_text + Colors.ENDC
            )
            output.append("\n\n\n")
            colored_output.append("\n\n\n")
            i += 1

        games_played_header = "\n#####AMOUNT OF GAMES PLAYED#####"
        output.append(games_played_header)
        colored_output.append(
            Colors.HEADER + Colors.BOLD + games_played_header + Colors.ENDC
        )

        for player in self.players:
            player_stats_text = f"{player.name} played {player.games_played} games, happiness: {np.round(player.happiness, 2)}"
            output.append(player_stats_text)

            # Color based on happiness level
            happiness = player.happiness
            if happiness >= np.mean([p.happiness for p in self.players]) + np.std(
                [p.happiness for p in self.players]
            ):
                color = Colors.OKGREEN  # High happiness
            elif happiness <= np.mean([p.happiness for p in self.players]) - np.std(
                [p.happiness for p in self.players]
            ):
                color = Colors.FAIL  # Low happiness
            else:
                color = Colors.ENDC  # Normal happiness

            colored_player_stats = f"{Colors.BOLD}{player.name}{Colors.ENDC} played {player.games_played} games, happiness: {color}{np.round(player.happiness, 2)}{Colors.ENDC}"
            colored_output.append(colored_player_stats)
        #####ADDD THE COUNT OF PLAYERS THAT PLAYED TOGETHER AT LEAST TWICE AND SO ON, REPLACE FROM PRINT_ALL_RESULTS#####
        player_pairs, pair_rounds = self.count_all_pairs(order_num_list)

        output, _ = self.add_team_repetition_to_output(
            output, player_pairs, pair_rounds, minimum=minimum_team_repeats
        )
        # Also add to colored output
        temp_output = []
        temp_output, _ = self.add_team_repetition_to_output(
            temp_output, player_pairs, pair_rounds, minimum=minimum_team_repeats
        )
        for line in temp_output:
            if line.startswith("\n#######PLAYERS WHO PLAYED TOGETHER"):
                colored_output.append(Colors.WARNING + Colors.BOLD + line + Colors.ENDC)
            else:
                colored_output.append(Colors.WARNING + line + Colors.ENDC)

        # Find and display players who played against each other at least twice and record the rounds
        opponent_pairs, opponent_pair_rounds = self.count_all_opponent_pairs(
            order_num_list
        )

        output, _ = self.add_opponent_team_repetition_to_output(
            output,
            opponent_pairs,
            opponent_pair_rounds,
            minimum=minimum_opponent_repeats,
        )
        # Also add to colored output
        temp_output = []
        temp_output, _ = self.add_opponent_team_repetition_to_output(
            temp_output,
            opponent_pairs,
            opponent_pair_rounds,
            minimum=minimum_opponent_repeats,
        )
        for line in temp_output:
            if line.startswith("\n#######PLAYERS WHO PLAYED AGAINST"):
                colored_output.append(Colors.FAIL + Colors.BOLD + line + Colors.ENDC)
            else:
                colored_output.append(Colors.FAIL + line + Colors.ENDC)

        # Add players who never met
        output, _ = self.add_never_met_pairs_to_output(output, order_num_list)
        temp_output = []
        temp_output, _ = self.add_never_met_pairs_to_output(temp_output, order_num_list)
        for line in temp_output:
            if line.startswith("\n#######PLAYERS THAT NEVER MET"):
                colored_output.append(Colors.OKCYAN + Colors.BOLD + line + Colors.ENDC)
            else:
                colored_output.append(Colors.OKCYAN + line + Colors.ENDC)

        # Add happiness analytics section
        analytics_header = "\n#####HAPPINESS ANALYTICS#####"
        output.append(analytics_header)
        colored_output.append(
            Colors.HEADER + Colors.BOLD + analytics_header + Colors.ENDC
        )

        # Calculate happiness statistics
        happiness_values = [player.happiness for player in self.players]

        avg_text = f"Average happiness: {np.round(np.mean(happiness_values), 2)}"
        output.append(avg_text)
        colored_output.append(Colors.OKCYAN + avg_text + Colors.ENDC)

        std_text = (
            f"Happiness standard deviation: {np.round(np.std(happiness_values), 2)}"
        )
        output.append(std_text)
        colored_output.append(Colors.OKCYAN + std_text + Colors.ENDC)

        min_text = f"Min happiness: {np.round(min(happiness_values), 2)}"
        output.append(min_text)
        colored_output.append(Colors.FAIL + min_text + Colors.ENDC)

        max_text = f"Max happiness: {np.round(max(happiness_values), 2)}"
        output.append(max_text)
        colored_output.append(Colors.OKGREEN + max_text + Colors.ENDC)

        # Identify happiest and least happy players
        happiest_players = sorted(
            self.players, key=lambda p: p.happiness, reverse=True
        )[:3]
        least_happy_players = sorted(self.players, key=lambda p: p.happiness)[:3]

        happiest_header = "\nHappiest players:"
        output.append(happiest_header)
        colored_output.append(
            Colors.OKGREEN + Colors.BOLD + happiest_header + Colors.ENDC
        )

        for player in happiest_players:
            player_happy_text = f"{player.name}: {np.round(player.happiness, 2)}"
            output.append(player_happy_text)
            colored_output.append(Colors.OKGREEN + player_happy_text + Colors.ENDC)

        least_happy_header = "\nLeast happy players:"
        output.append(least_happy_header)
        colored_output.append(
            Colors.FAIL + Colors.BOLD + least_happy_header + Colors.ENDC
        )

        for player in least_happy_players:
            player_sad_text = f"{player.name}: {np.round(player.happiness, 2)}"
            output.append(player_sad_text)
            colored_output.append(Colors.FAIL + player_sad_text + Colors.ENDC)

        print("\n".join(colored_output))  # Print colored version to terminal

    def export_to_excel(
        self,
        directory=None,
        date_str=datetime.datetime.now().strftime("%d_%m_%Y"),
        filename=None,
    ):
        """
        Export all games from the session to an Excel file.

        Parameters:
        - directory: Directory to save the file (default: "sessions")
        - filename: Name of the file (default: "session_{date in dd%mm%yy format}.xlsx")

        Returns:
        - str: Path to the saved file
        """

        # Create directory if it doesn't exist
        if directory is None:
            directory = os.getcwd()

        if not os.path.exists(directory):
            os.makedirs(directory)

        # Generate filename if not provided
        if filename is None:
            # Create a folder with date
            filename = f"session_{date_str}.xlsx"

        # Ensure .xlsx extension
        if not filename.endswith(".xlsx"):
            filename += ".xlsx"

        filepath = os.path.join(directory, filename)

        # Create Excel writer
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            # Remove default sheet if it exists
            if "Sheet" in writer.book.sheetnames:
                del writer.book["Sheet"]

            # Sheet 1: Games by Round with formatted tables
            ws_games = writer.book.create_sheet("Games", 0)

            # Define styles
            header_fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            header_font = Font(bold=True, color="FFFFFF", size=11)
            team_a_fill = PatternFill(
                start_color="B4C7E7", end_color="B4C7E7", fill_type="solid"
            )
            team_b_fill = PatternFill(
                start_color="F4B084", end_color="F4B084", fill_type="solid"
            )
            vs_fill = PatternFill(
                start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"
            )
            level_fill = PatternFill(
                start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
            )
            round_title_fill = PatternFill(
                start_color="305496", end_color="305496", fill_type="solid"
            )
            round_title_font = Font(bold=True, color="FFFFFF", size=14)
            center_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            current_row = 1

            for round_idx, round in enumerate(self.rounds, start=1):
                # Round title
                ws_games.merge_cells(f"A{current_row}:H{current_row}")
                cell = ws_games[f"A{current_row}"]
                cell.value = f"ROUND {round_idx}"
                cell.fill = round_title_fill
                cell.font = round_title_font
                cell.alignment = center_alignment
                current_row += 1

                # Headers
                headers = [
                    "Game",
                    "Team A Player 1",
                    "Team A Player 2",
                    "VS",
                    "Team B Player 1",
                    "Team B Player 2",
                    "Level Diff",
                    "Info",
                ]
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws_games.cell(row=current_row, column=col_idx)
                    cell.value = header
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_alignment
                    cell.border = thin_border

                current_row += 1
                start_data_row = current_row

                # Game data
                for game_idx, game in enumerate(round.games, start=1):
                    team_A, team_B = list(game.teams)
                    team_A_players = [p.name for p in team_A.players]
                    team_B_players = [p.name for p in team_B.players]

                    # Game number
                    cell = ws_games.cell(row=current_row, column=1)
                    cell.value = game_idx
                    cell.alignment = center_alignment
                    cell.border = thin_border

                    # Team A players
                    for i, col in enumerate([2, 3]):
                        cell = ws_games.cell(row=current_row, column=col)
                        cell.value = (
                            team_A_players[i] if i < len(team_A_players) else ""
                        )
                        cell.fill = team_a_fill
                        cell.alignment = center_alignment
                        cell.border = thin_border

                    # VS
                    cell = ws_games.cell(row=current_row, column=4)
                    cell.value = "VS"
                    cell.fill = vs_fill
                    cell.alignment = center_alignment
                    cell.border = thin_border
                    cell.font = Font(bold=True)

                    # Team B players
                    for i, col in enumerate([5, 6]):
                        cell = ws_games.cell(row=current_row, column=col)
                        cell.value = (
                            team_B_players[i] if i < len(team_B_players) else ""
                        )
                        cell.fill = team_b_fill
                        cell.alignment = center_alignment
                        cell.border = thin_border

                    # Level difference
                    cell = ws_games.cell(row=current_row, column=7)
                    cell.value = np.round(game.level_difference, 2)
                    cell.fill = level_fill
                    cell.alignment = center_alignment
                    cell.border = thin_border

                    # Info (type/gender preference)
                    cell = ws_games.cell(row=current_row, column=8)
                    info_text = (
                        str(game.type_preference) if game.type_preference else ""
                    )
                    if game.gender_preference:
                        info_text += f" / {game.gender_preference}"
                    cell.value = info_text
                    cell.alignment = center_alignment
                    cell.border = thin_border

                    current_row += 1

                # Not playing section for this round
                if round.not_playing:
                    current_row += 1
                    ws_games.merge_cells(f"A{current_row}:H{current_row}")
                    cell = ws_games[f"A{current_row}"]
                    not_playing_names = ", ".join([p.name for p in round.not_playing])
                    cell.value = f"Not Playing: {not_playing_names}"
                    cell.fill = PatternFill(
                        start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"
                    )
                    cell.font = Font(italic=True)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.border = thin_border
                    current_row += 1

                # Add spacing between rounds
                current_row += 2

            # Adjust column widths
            column_widths = [8, 18, 18, 6, 18, 18, 12, 20]
            for col_idx, width in enumerate(column_widths, start=1):
                ws_games.column_dimensions[get_column_letter(col_idx)].width = width

            # Sheet 2: Player Statistics (sorted alphabetically)
            player_stats = []
            for player in self.players:
                player_stats.append(
                    {
                        "Player Name": player.name,
                        "Level": player.level,
                        "Games Played": player.games_played,
                        "Final Happiness": np.round(player.happiness, 2),
                        "Gender": getattr(player, "gender", "Unknown"),
                    }
                )

            df_players = pd.DataFrame(player_stats)
            df_players = df_players.sort_values("Player Name", ascending=True)
            df_players.to_excel(writer, sheet_name="Player Stats", index=False)

            # Sheet 3: Session Variables
            session_variables = {
                "Variable": [
                    "amount_of_rounds",
                    "games_per_round_each_round",
                    "players_per_team_each_round",
                    "type_preferences",
                    "gender_preferences",
                    "level_gap_tol",
                    "num_iter",
                    "spectrum",
                    "weight_same_teammate",
                    "rounds_reordering",
                    "mean_happiness",
                    "std_happiness",
                    "max_happiness",
                    "min_happiness",
                    "max_happiness_difference",
                ],
                "Value": [
                    self.amount_of_rounds,
                    str(self.games_per_round_each_round),
                    str(self.players_per_team_each_round),
                    str(self.type_preferences),
                    str(self.gender_preferences),
                    self.level_gap_tol,
                    self.num_iter,
                    self.spectrum,
                    self.weight_same_teammate,
                    str(self.rounds_reordering),
                    np.round(self.mean_happiness, 2),
                    np.round(self.std_happiness, 2),
                    np.round(self.max_and_min_happiness[0], 2),
                    np.round(self.max_and_min_happiness[1], 2),
                    np.round(self.max_happiness_difference, 2),
                ],
            }
            df_session_vars = pd.DataFrame(session_variables)
            df_session_vars.to_excel(
                writer, sheet_name="Session Variables", index=False
            )

            # Sheet 4: Teammate Repetitions
            player_pairs, pair_rounds = self.count_all_pairs()
            repetitions_data = []
            for pair, count in player_pairs.items():
                if count >= 2:
                    pair_list = list(pair)
                    repetitions_data.append(
                        {
                            "Player 1": pair_list[0],
                            "Player 2": pair_list[1],
                            "Times Played Together": count,
                            "Rounds": ", ".join(map(str, pair_rounds[pair])),
                        }
                    )

            if repetitions_data:
                df_repetitions = pd.DataFrame(repetitions_data)
                df_repetitions = df_repetitions.sort_values(
                    "Times Played Together", ascending=False
                )
                df_repetitions.to_excel(
                    writer, sheet_name="Teammate Repetitions", index=False
                )

            # Sheet 5: Opponent Repetitions
            opponent_pairs, opponent_pair_rounds = self.count_all_opponent_pairs()
            opponent_repetitions_data = []
            for pair, count in opponent_pairs.items():
                if count >= 2:
                    pair_list = list(pair)
                    opponent_repetitions_data.append(
                        {
                            "Player 1": pair_list[0],
                            "Player 2": pair_list[1],
                            "Times Played Against": count,
                            "Rounds": ", ".join(map(str, opponent_pair_rounds[pair])),
                        }
                    )

            if opponent_repetitions_data:
                df_opponent_reps = pd.DataFrame(opponent_repetitions_data)
                df_opponent_reps = df_opponent_reps.sort_values(
                    "Times Played Against", ascending=False
                )
                df_opponent_reps.to_excel(
                    writer, sheet_name="Opponent Repetitions", index=False
                )

            # Sheet 6: Session Summary
            summary_data = [
                {
                    "Total Rounds": self.amount_of_rounds,
                    "Total Players": len(self.players),
                    "Mean Happiness": np.round(self.mean_happiness, 2),
                    "Std Happiness": np.round(self.std_happiness, 2),
                    "Min Happiness": np.round(self.max_and_min_happiness[1], 2),
                    "Max Happiness": np.round(self.max_and_min_happiness[0], 2),
                    "Happiness Range": np.round(self.max_happiness_difference, 2),
                    "Weight Same Teammate": self.weight_same_teammate,
                    "Level Gap Tolerance": self.level_gap_tol,
                }
            ]

            df_summary = pd.DataFrame(summary_data)
            df_summary = df_summary.T
            df_summary.columns = ["Value"]
            df_summary.to_excel(writer, sheet_name="Summary")

        # Save the changeable version
        print(f"Session exported successfully to: {filepath}")

        # Create read-only version
        readonly_filename = f"session_{date_str}_read_only.xlsx"
        readonly_filepath = os.path.join(directory, readonly_filename)

        # Copy the workbook for read-only version
        import shutil

        shutil.copy(filepath, readonly_filepath)

        # Set workbook security to make it read-only
        from openpyxl import load_workbook

        wb_readonly = load_workbook(readonly_filepath)

        # Protect the workbook structure and windows
        wb_readonly.security.lockStructure = True
        wb_readonly.security.lockWindows = True
        wb_readonly.security.workbookPassword = ""

        # Protect each sheet
        for sheet in wb_readonly.worksheets:
            sheet.protection.sheet = True
            sheet.protection.password = ""
            sheet.protection.enable()

        wb_readonly.save(readonly_filepath)
        print(f"Read-only version saved to: {readonly_filepath}")

        return filepath

    def save_session_of_rounds(
        self,
        date_str=datetime.datetime.now().strftime("%d_%m_%Y"),
        main_folder="sessions",
        export_to_excel=True,
        create_plots=True,
        **kwargs,
    ):

        # Create the session folder path
        session_folder = os.path.join(main_folder, date_str)
        # Create the folder if it doesn't exist, or find an available suffix
        counter = 2
        while os.path.exists(session_folder):
            session_folder = os.path.join(main_folder, f"{date_str}_{counter}")
            counter += 1
        os.makedirs(session_folder, exist_ok=True)
        from core.pickle_helper import save_session, load_session
        from core.charts import create_all_session_charts

        # Save
        save_session(
            self,
            folder=session_folder,
            filename=f"session_of_rounds_{date_str}.pkl",
        )
        if export_to_excel:
            self.export_to_excel(directory=session_folder, date_str=date_str, **kwargs)

        if create_plots:
            plots_dir = os.path.join(session_folder, "plots")
            create_all_session_charts(self, save_png=True, png_dir=plots_dir)

    def print_happiness_with_colors(self):
        """
        Print all players' happiness sorted from least happy first, with color coding:
        - Red: First quartile (least happy 25%)
        - Yellow: Second and third quartiles (middle 50%)
        - Green: Fourth quartile (happiest 25%)
        """
        # Get happiness values and sort players by happiness
        sorted_players = sorted(self.players, key=lambda p: p.happiness)
        n = len(sorted_players)

        # Calculate quartile boundaries
        q1_idx = n // 4
        q3_idx = 3 * n // 4

        print("\n" + "=" * 80)
        print("ALL PLAYERS HAPPINESS (Least Happy → Most Happy)")
        print("=" * 80)

        for idx, player in enumerate(sorted_players):
            # Determine color based on quartile
            if idx < q1_idx:
                # Red for first quartile (least happy)
                color_code = "\033[91m"  # Red
            elif idx >= q3_idx:
                # Green for fourth quartile (most happy)
                color_code = "\033[92m"  # Green
            else:
                # Yellow for second and third quartiles (median)
                color_code = "\033[93m"  # Yellow

            reset_code = "\033[0m"
            happiness_str = f"{player.happiness:.2f}"
            print(f"{color_code}{player.name:20s} → {happiness_str:>8s}{reset_code}")

        print("=" * 80 + "\n")

    def recalculate_session_statistics(self):
        """
        Recalculate session-level statistics (mean and std happiness).
        Call this after making changes to rounds/games/players.
        """
        self.mean_happiness = np.mean([player.happiness for player in self.players])
        self.std_happiness = np.std([player.happiness for player in self.players])

        # Also update relative happiness for each player
        for player in self.players:
            player.relative_happiness = player.happiness - self.mean_happiness

        # Update max/min happiness and difference
        self.max_and_min_happiness = (
            np.max([player.happiness for player in self.players]),
            np.min([player.happiness for player in self.players]),
        )
        self.max_happiness_difference = (
            self.max_and_min_happiness[0] - self.max_and_min_happiness[1]
        )

        # Update least and most happy players
        self.least_happy_players = sorted(self.players, key=lambda p: p.happiness)[:3]
        self.happiest_players = sorted(
            self.players, key=lambda p: p.happiness, reverse=True
        )[:3]

    def apply_changes_to_rounds(self, modified_round_indices):
        """
        Apply changes to multiple rounds and recalculate all statistics.

        Parameters:
        -----------
        modified_round_indices : list of int
            List of round indices that have been modified
        """
        if not modified_round_indices:
            return

        # Find the earliest modified round
        min_round_idx = min(modified_round_indices)

        # Do not globally subtract happiness here.
        # Each round's recalculate_happiness() already rolls back that round's prior gain
        # before applying the new one. Doing both causes double subtraction and large drops.

        # Now recalculate happiness for all rounds from min_round_idx onwards
        print(f"Recalculating happiness from Round {min_round_idx + 1} onwards...")
        for round_idx in range(min_round_idx, len(self.rounds)):
            self.rounds[round_idx].recalculate_happiness(round_idx=round_idx)

        self.recalculate_session_statistics()

    # example usage:
    # current_date = datetime.datetime.now().strftime("%d_%m_%Y")
    # save_session_of_rounds(session_of_rounds, current_date)
    # Load:
    # session_of_rounds = load_session(f"sessions/{current_date}/session_of_rounds_{current_date}.pkl")


# %%
################################################################################
################################################################################
################################################################################
###############################                 ################################
###############################  #####  #   #  ################################
############################### #        ###   ################################
###############################  ####     #    ################################
###############################      #    #    ################################
############################### #####     #    ################################
###############################                 ################################
################################################################################
################################################################################
################################################################################
